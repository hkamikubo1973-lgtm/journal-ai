from datetime import datetime
import inspect
import io
from pathlib import Path
import sys
import tempfile
import unittest
from unittest.mock import patch

import pandas as pd
from fastapi import HTTPException
from openpyxl import load_workbook


SRC_DIR = Path(__file__).resolve().parent.parent / "src"
sys.path.insert(0, str(SRC_DIR))

from api.journal import (  # noqa: E402
    InputExcelRequest,
    post_export_input_excel,
    post_save_input_excel,
)
from columns import EPSON_COLUMNS  # noqa: E402
from export_file_service import ExportFileError  # noqa: E402
import input_excel_save_service  # noqa: E402
import input_excel_service  # noqa: E402
from input_excel_save_service import (  # noqa: E402
    INPUT_EXCEL_EXPORT_SUBDIR,
    InputExcelSaveError,
    InputExcelSaveResult,
    save_input_excel,
)
from input_excel_service import (  # noqa: E402
    INPUT_CSV_COLUMNS,
    InputExcelExport,
    InputExcelValidationError,
    build_input_csv_rows,
    build_input_excel_rows,
    build_input_journal_excel,
    export_input_excel,
)
from journal_registration_service import (  # noqa: E402
    EDIT_FORM_FIELDS,
    build_registration_id,
)


FIXED_DATETIME = datetime(2026, 8, 21, 12, 34)
COLUMN_WIDTHS = [4, 10, 12, 14, 11, 12, 14, 11, 24, 18, 8, 18]


class InputExcelServiceTest(unittest.TestCase):
    def test_single_item_has_fixed_header_and_one_row(self):
        result = export_input_excel(
            [self._item("A")],
            export_datetime=FIXED_DATETIME,
        )
        worksheet = self._worksheet(result.content)

        self.assertEqual(
            result.filename,
            "input_journal_print_20260821_1234.xlsx",
        )
        self.assertEqual(
            [cell.value for cell in worksheet[1]],
            INPUT_CSV_COLUMNS,
        )
        self.assertEqual(worksheet.max_column, 12)
        self.assertEqual(worksheet.max_row, 2)

    def test_multiple_items_keep_cart_order_and_number_consecutively(self):
        result = export_input_excel(
            [self._item("C"), self._item("A"), self._item("B")],
            export_datetime=FIXED_DATETIME,
        )
        worksheet = self._worksheet(result.content)

        self.assertEqual(
            [worksheet.cell(row=row, column=1).value for row in range(2, 5)],
            [1, 2, 3],
        )
        self.assertEqual(
            [worksheet.cell(row=row, column=9).value for row in range(2, 5)],
            ["摘要C", "摘要A", "摘要B"],
        )

    def test_print_values_date_and_amount_are_mapped_by_backend(self):
        item = self._item("A")
        item["print_metadata"]["print_category"] = "通常"
        item["print_warnings"] = ["DB雛形なし", "伝票摘要なし"]
        item["prepared_journal"]["amount"] = "1,234"
        item["epson_base_row"]["借方金額"] = "1,234"
        item["epson_base_row"]["貸方金額"] = "1,234"
        item["registration_id"] = build_registration_id(
            item["prepared_journal"],
            item["epson_base_row"],
        )

        result = export_input_excel([item], export_datetime=FIXED_DATETIME)
        worksheet = self._worksheet(result.content)

        self.assertEqual(worksheet["B2"].value, "20260821")
        self.assertEqual(worksheet["E2"].value, 1234)
        self.assertEqual(worksheet["H2"].value, 1234)
        self.assertEqual(worksheet["E2"].number_format, "#,##0")
        self.assertEqual(worksheet["H2"].number_format, "#,##0")
        self.assertEqual(worksheet["K2"].value, "通常")
        self.assertEqual(
            worksheet["L2"].value,
            "DB雛形なし / 伝票摘要なし",
        )

    def test_workbook_format_matches_streamlit_specification(self):
        worksheet = self._worksheet(
            export_input_excel(
                [self._item("A")],
                export_datetime=FIXED_DATETIME,
            ).content
        )

        self.assertEqual(worksheet.title, "簡易仕訳帳")
        self.assertEqual(
            [
                worksheet.column_dimensions[letter].width
                for letter in "ABCDEFGHIJKL"
            ],
            COLUMN_WIDTHS,
        )
        self.assertEqual(worksheet.freeze_panes, "A2")
        self.assertEqual(worksheet.auto_filter.ref, "A1:L2")
        self.assertEqual(worksheet.print_title_rows, "$1:$1")
        self.assertEqual(worksheet.page_setup.paperSize, 9)
        self.assertEqual(worksheet.page_setup.orientation, "landscape")
        self.assertEqual(worksheet.page_setup.fitToWidth, 1)
        self.assertEqual(worksheet.page_setup.fitToHeight, 0)
        self.assertTrue(worksheet.sheet_properties.pageSetUpPr.fitToPage)
        self.assertEqual(worksheet.page_margins.left, 0.25)
        self.assertEqual(worksheet.page_margins.right, 0.25)
        self.assertEqual(worksheet.page_margins.top, 0.5)
        self.assertEqual(worksheet.page_margins.bottom, 0.5)
        self.assertEqual(worksheet.page_margins.header, 0.2)
        self.assertEqual(worksheet.page_margins.footer, 0.2)
        self.assertTrue(worksheet["A1"].font.bold)
        self.assertEqual(worksheet["A1"].fill.fgColor.rgb, "00D9EAF7")
        self.assertEqual(worksheet["A1"].alignment.horizontal, "center")
        self.assertEqual(worksheet["A1"].alignment.vertical, "center")
        self.assertTrue(worksheet["A1"].alignment.wrap_text)
        self.assertEqual(worksheet["A2"].alignment.vertical, "top")
        self.assertFalse(worksheet["A2"].alignment.wrap_text)
        for cell in (worksheet["I2"], worksheet["J2"], worksheet["L2"]):
            self.assertTrue(cell.alignment.wrap_text)
        for cell in (worksheet["A1"], worksheet["L2"]):
            self.assertEqual(cell.border.left.style, "thin")
            self.assertEqual(cell.border.left.color.rgb, "00B7B7B7")

    def test_empty_cart_is_rejected_without_workbook(self):
        with patch("input_excel_service.build_input_journal_excel") as builder:
            with self.assertRaisesRegex(
                InputExcelValidationError,
                "出力対象がありません",
            ):
                export_input_excel([], export_datetime=FIXED_DATETIME)

        builder.assert_not_called()

    def test_tampered_registration_id_is_rejected(self):
        item = self._item("A")
        item["registration_id"] = "tampered"

        with self.assertRaises(InputExcelValidationError):
            export_input_excel([item], export_datetime=FIXED_DATETIME)

    def test_missing_base_column_is_rejected(self):
        item = self._item("A")
        del item["epson_base_row"]["借方消費税コード"]

        with self.assertRaises(InputExcelValidationError):
            export_input_excel([item], export_datetime=FIXED_DATETIME)

    def test_print_metadata_and_warnings_are_strictly_validated(self):
        missing_metadata = self._item("A")
        del missing_metadata["print_metadata"]
        invalid_category = self._item("B")
        invalid_category["print_metadata"]["print_category"] = 1
        invalid_warnings = self._item("C")
        invalid_warnings["print_warnings"] = ["注意", 1]

        for item in (missing_metadata, invalid_category, invalid_warnings):
            with self.subTest(marker=item["prepared_journal"]["summary"]):
                with self.assertRaises(InputExcelValidationError):
                    build_input_excel_rows([item])

    def test_download_route_returns_raw_xlsx_and_filename(self):
        request = InputExcelRequest(items=[self._item("A")])
        expected = InputExcelExport(
            content=b"raw-xlsx-bytes",
            filename="input_journal_print_20260821_1234.xlsx",
        )

        with patch("api.journal.export_input_excel", return_value=expected):
            response = post_export_input_excel(request)

        self.assertEqual(response.body, expected.content)
        self.assertEqual(
            response.headers["content-disposition"],
            'attachment; filename="input_journal_print_20260821_1234.xlsx"',
        )
        self.assertEqual(
            response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_routes_reject_empty_cart_and_tampered_id(self):
        for items in ([], [self._tampered_item()]):
            with self.subTest(item_count=len(items)):
                request = InputExcelRequest(items=items)
                for route in (post_export_input_excel, post_save_input_excel):
                    with self.subTest(route=route.__name__):
                        with self.assertRaises(HTTPException) as raised:
                            route(request)
                        self.assertEqual(raised.exception.status_code, 422)

    def test_save_creates_02_subdir_and_writes_download_bytes(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            base_dir = Path(temp_dir) / "base"
            base_dir.mkdir()
            generated = export_input_excel(
                [self._item("A")],
                export_datetime=FIXED_DATETIME,
            )

            result = save_input_excel(
                [self._item("A")],
                export_dir=str(base_dir),
                export_datetime=FIXED_DATETIME,
                export_builder=lambda items, export_datetime=None: generated,
            )

            saved_path = Path(result.saved_path)
            self.assertTrue(result.success)
            self.assertEqual(saved_path.parent.name, INPUT_EXCEL_EXPORT_SUBDIR)
            self.assertEqual(saved_path.read_bytes(), generated.content)
            self.assertIn("検索DBは更新していません", result.message)

    def test_missing_base_dir_does_not_fallback(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            missing_dir = Path(temp_dir) / "missing"

            with self.assertRaises(InputExcelSaveError):
                save_input_excel(
                    [self._item("A")],
                    export_dir=str(missing_dir),
                    export_datetime=FIXED_DATETIME,
                )

            self.assertFalse(missing_dir.exists())

    def test_save_failure_reports_no_db_update(self):
        with self.assertRaisesRegex(
            InputExcelSaveError,
            "検索DBは更新していません",
        ):
            save_input_excel(
                [self._item("A")],
                export_dir="X:/unused",
                export_datetime=FIXED_DATETIME,
                file_saver=lambda *args: (_ for _ in ()).throw(
                    ExportFileError("書込み不能")
                ),
            )

    def test_save_route_returns_structured_result(self):
        request = InputExcelRequest(items=[self._item("A")])
        expected = InputExcelSaveResult(
            success=True,
            filename="input_journal_print_20260821_1234.xlsx",
            saved_path="X:/02_入力用Excel/input.xlsx",
            message="入力用Excelを保存しました。検索DBは更新していません。",
        )

        with patch("api.journal.save_input_excel", return_value=expected):
            response = post_save_input_excel(request)

        self.assertEqual(response, expected.to_dict())

    def test_streamlit_and_cart_paths_are_semantically_identical(self):
        item = self._item("互換")
        item["prepared_journal"]["voucher_summary"] = ""
        item["epson_base_row"]["伝票摘要"] = ""
        item["print_metadata"]["print_category"] = "通常"
        item["print_warnings"] = ["DB雛形なし", "伝票摘要なし"]
        item["registration_id"] = build_registration_id(
            item["prepared_journal"],
            item["epson_base_row"],
        )
        legacy_source = dict(item["epson_base_row"])
        legacy_source["区分"] = "通常"
        legacy_source["DB雛形"] = "なし"
        legacy_rows = build_input_csv_rows([[legacy_source]])
        legacy_frame = pd.DataFrame(
            legacy_rows,
            columns=INPUT_CSV_COLUMNS,
        ).fillna("")
        legacy_content = build_input_journal_excel(legacy_frame)
        cart_content = export_input_excel(
            [item],
            export_datetime=FIXED_DATETIME,
        ).content

        self.assertEqual(
            self._workbook_snapshot(legacy_content),
            self._workbook_snapshot(cart_content),
        )

    def test_input_excel_modules_have_no_db_update_dependency(self):
        sources = (
            inspect.getsource(input_excel_service),
            inspect.getsource(input_excel_save_service),
        )
        for source in sources:
            for forbidden_name in (
                "journal_persistence_service",
                "register_epson_rows_to_search_db",
                "update_search_csv",
                "append_to_csv",
                "keep_recent_years",
                "TRANSACTIONS_PATH",
                "transactions.csv",
            ):
                self.assertNotIn(forbidden_name, source)

    @staticmethod
    def _worksheet(content):
        return load_workbook(io.BytesIO(content)).active

    @classmethod
    def _workbook_snapshot(cls, content):
        worksheet = cls._worksheet(content)
        return {
            "title": worksheet.title,
            "values": [
                [cell.value for cell in row]
                for row in worksheet.iter_rows()
            ],
            "widths": [
                worksheet.column_dimensions[letter].width
                for letter in "ABCDEFGHIJKL"
            ],
            "freeze": worksheet.freeze_panes,
            "filter": worksheet.auto_filter.ref,
            "print_titles": worksheet.print_title_rows,
            "paper_size": worksheet.page_setup.paperSize,
            "orientation": worksheet.page_setup.orientation,
            "fit_width": worksheet.page_setup.fitToWidth,
            "fit_height": worksheet.page_setup.fitToHeight,
            "fit_to_page": worksheet.sheet_properties.pageSetUpPr.fitToPage,
            "margins": (
                worksheet.page_margins.left,
                worksheet.page_margins.right,
                worksheet.page_margins.top,
                worksheet.page_margins.bottom,
                worksheet.page_margins.header,
                worksheet.page_margins.footer,
            ),
            "styles": [
                [
                    (
                        cell.font.bold,
                        cell.fill.fgColor.rgb,
                        cell.alignment.horizontal,
                        cell.alignment.vertical,
                        cell.alignment.wrap_text,
                        cell.border.left.style,
                        cell.border.left.color.rgb,
                        cell.number_format,
                    )
                    for cell in row
                ]
                for row in worksheet.iter_rows()
            ],
        }

    @staticmethod
    def _prepared(marker):
        values = {
            "voucher_date": "20260821",
            "voucher_no": f"CERT-{marker}",
            "voucher_summary": f"伝票摘要{marker}",
            "debit_account_code": "100",
            "debit_account_name": "現金",
            "debit_sub_code": "1",
            "debit_sub_name": "小口",
            "debit_dept_code": "10",
            "debit_dept_name": "営業",
            "credit_account_code": "200",
            "credit_account_name": "売上",
            "credit_sub_code": "",
            "credit_sub_name": "",
            "credit_dept_code": "20",
            "credit_dept_name": "管理",
            "amount": 1234,
            "summary": f"摘要{marker}",
            "source_debit_amount": "1234",
            "source_credit_amount": "1234",
        }
        return {field: values[field] for field in EDIT_FORM_FIELDS}

    @classmethod
    def _item(cls, marker):
        prepared = cls._prepared(marker)
        base_row = {column: f"元値:{column}" for column in EPSON_COLUMNS}
        base_row.update({
            "伝票日付": prepared["voucher_date"],
            "証番号": prepared["voucher_no"],
            "伝票摘要": prepared["voucher_summary"],
            "借方部門": prepared["debit_dept_code"],
            "借方部門名": prepared["debit_dept_name"],
            "借方科目": prepared["debit_account_code"],
            "借方科目名": prepared["debit_account_name"],
            "借方補助": prepared["debit_sub_code"],
            "借方補助科目名": prepared["debit_sub_name"],
            "借方金額": str(prepared["amount"]),
            "貸方部門": prepared["credit_dept_code"],
            "貸方部門名": prepared["credit_dept_name"],
            "貸方科目": prepared["credit_account_code"],
            "貸方科目名": prepared["credit_account_name"],
            "貸方補助": prepared["credit_sub_code"],
            "貸方補助科目名": prepared["credit_sub_name"],
            "貸方金額": str(prepared["amount"]),
            "摘要": prepared["summary"],
        })
        return {
            "registration_id": build_registration_id(prepared, base_row),
            "prepared_journal": prepared,
            "epson_base_row": base_row,
            "print_metadata": {"print_category": ""},
            "print_warnings": [],
        }

    @classmethod
    def _tampered_item(cls):
        item = cls._item("TAMPERED")
        item["registration_id"] = "tampered"
        return item


if __name__ == "__main__":
    unittest.main()
