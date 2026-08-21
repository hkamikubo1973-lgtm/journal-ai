from datetime import date, datetime
import hashlib
import inspect
import io
from pathlib import Path
import sys
import unittest

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parent.parent
SRC_DIR = ROOT_DIR / "src"
sys.path.insert(0, str(SRC_DIR))

import receivable_report_service  # noqa: E402
from receivable_report_service import (  # noqa: E402
    RECEIVABLE_CHECK_COLUMNS,
    build_receivable_check_excel,
    build_receivable_check_filename,
    build_receivable_check_rows,
    format_receivable_check_amount,
    format_receivable_check_date,
)


COLUMN_WIDTHS = [4, 11, 16, 11, 11, 10, 18, 22, 20]


class ReceivableReportServiceTest(unittest.TestCase):
    def test_single_settlement_creates_one_fixed_nine_column_row(self):
        rows = build_receivable_check_rows([self._journal("A")])

        self.assertEqual(len(rows), 1)
        self.assertEqual(list(rows[0]), RECEIVABLE_CHECK_COLUMNS)
        self.assertEqual(len(RECEIVABLE_CHECK_COLUMNS), 9)
        self.assertEqual(rows[0]["No"], 1)

    def test_multiple_settlements_keep_input_order_and_number_consecutively(self):
        rows = build_receivable_check_rows([
            self._journal("C"),
            self._journal("A"),
            self._journal("B"),
        ])

        self.assertEqual([row["取引先"] for row in rows], ["C社", "A社", "B社"])
        self.assertEqual([row["No"] for row in rows], [1, 2, 3])

    def test_multiple_fifo_candidates_stay_in_one_row_with_line_breaks(self):
        journal = self._journal("A")
        journal["source_candidates"] = [
            {"未収科目": "未収運賃", "消込予定": 100000, "摘要": "1便"},
            {"未収科目": "売掛金", "消込予定": "115,000", "摘要": "2便"},
        ]

        rows = build_receivable_check_rows([journal])

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["未収内容"], "未収運賃 100,000\n売掛金 115,000")

    def test_multiple_generated_journals_use_line_breaks(self):
        journal = self._journal("A")
        journal["rows"] = [
            {"借方科目": "普通預金", "貸方科目": "未収運賃", "金額": 210000, "摘要": "入金"},
            {"借方科目": "普通預金", "貸方科目": "仮受金", "金額": 5000, "摘要": "過入金"},
        ]

        row = build_receivable_check_rows([journal])[0]

        self.assertEqual(
            row["生成仕訳"],
            "普通預金 / 未収運賃 210,000\n普通預金 / 仮受金 5,000",
        )

    def test_generated_journal_summaries_are_deduplicated_in_first_seen_order(self):
        journal = self._journal("A")
        journal["rows"] = [
            {"摘要": "入金"},
            {"摘要": "入金"},
            {"摘要": "差額調整"},
        ]

        row = build_receivable_check_rows([journal])[0]

        self.assertEqual(row["摘要"], "入金\n差額調整")

    def test_candidate_summaries_are_fallback_when_journal_has_no_summary(self):
        journal = self._journal("A")
        journal["rows"] = [{"借方科目": "普通預金", "貸方科目": "未収運賃", "金額": 215000}]
        journal["source_candidates"] = [
            {"未収科目": "未収運賃", "消込予定": 100000, "摘要": "6月分"},
            {"未収科目": "未収運賃", "消込予定": 115000, "摘要": "7月分"},
        ]

        row = build_receivable_check_rows([journal])[0]

        self.assertEqual(row["摘要"], "6月分\n7月分")

    def test_payment_amount_uses_explicit_value(self):
        journal = self._journal("A")
        journal["payment_amount"] = "215,000"

        self.assertEqual(build_receivable_check_rows([journal])[0]["入金額"], 215000)

    def test_missing_payment_amount_uses_target_total_plus_difference(self):
        journal = self._journal("A")
        journal["payment_amount"] = ""
        journal["target_total"] = 200000
        journal["difference"] = 15000

        self.assertEqual(build_receivable_check_rows([journal])[0]["入金額"], 215000)

    def test_missing_payment_and_difference_use_journal_amount_total(self):
        journal = self._journal("A")
        journal["payment_amount"] = None
        journal["difference"] = None
        journal["rows"] = [
            {"金額": 100000},
            {"金額": "115,000"},
        ]

        self.assertEqual(build_receivable_check_rows([journal])[0]["入金額"], 215000)

    def test_missing_target_total_uses_candidate_scheduled_amount_total(self):
        journal = self._journal("A")
        journal["target_total"] = None
        journal["source_candidates"] = [
            {"消込予定": 100000},
            {"消込予定": "110,000"},
        ]

        self.assertEqual(build_receivable_check_rows([journal])[0]["消込額"], 210000)

    def test_missing_difference_uses_payment_minus_target(self):
        journal = self._journal("A")
        journal["payment_amount"] = 215000
        journal["target_total"] = 210000
        journal["difference"] = ""

        self.assertEqual(build_receivable_check_rows([journal])[0]["差額"], 5000)

    def test_date_and_amount_formatting_match_legacy_behavior(self):
        self.assertEqual(format_receivable_check_date(date(2026, 8, 21)), "2026/08/21")
        self.assertEqual(format_receivable_check_date(datetime(2026, 8, 21, 12, 34)), "2026/08/21")
        self.assertEqual(format_receivable_check_date("20260821"), "2026/08/21")
        self.assertEqual(format_receivable_check_date("2026-08-21"), "2026/08/21")
        self.assertEqual(format_receivable_check_date("不明"), "不明")
        self.assertEqual(format_receivable_check_amount("215,000"), "215,000")
        self.assertEqual(format_receivable_check_amount("invalid"), "0")

    def test_single_id_filename_uses_sanitized_first_eight_characters(self):
        journal = self._journal("A")
        journal["settlement_id"] = "ab!c123456789"
        journal["settlement_date"] = "2026-08-21"

        self.assertEqual(
            build_receivable_check_filename([journal]),
            "receivable_check_20260821_abc12345.xlsx",
        )

    def test_multiple_id_filename_sorts_ids_hashes_and_uses_oldest_date(self):
        journals = [
            self._journal("Z", settlement_id="z-id", settlement_date="2026/08/22"),
            self._journal("A", settlement_id="a-id", settlement_date="2026/08/20"),
            self._journal("A2", settlement_id="a-id", settlement_date="2026/08/21"),
        ]
        expected_hash = hashlib.sha256(b"a-id|z-id").hexdigest()[:8]

        self.assertEqual(
            build_receivable_check_filename(journals),
            f"receivable_check_20260820_{expected_hash}.xlsx",
        )

    def test_filename_uses_undated_when_no_valid_date_exists(self):
        journal = self._journal("A")
        journal["settlement_date"] = ""

        self.assertEqual(
            build_receivable_check_filename([journal]),
            "receivable_check_undated_aaaaaaaa.xlsx",
        )

    def test_workbook_has_legacy_sheet_headers_and_column_widths(self):
        worksheet = self._worksheet(self._legacy_fixture())

        self.assertEqual(worksheet.title, "未収消込確認表")
        self.assertEqual([cell.value for cell in worksheet[1]], RECEIVABLE_CHECK_COLUMNS)
        self.assertEqual(worksheet.max_column, 9)
        self.assertEqual(
            [worksheet.column_dimensions[letter].width for letter in "ABCDEFGHI"],
            COLUMN_WIDTHS,
        )

    def test_workbook_header_border_and_wrap_match_legacy_style(self):
        worksheet = self._worksheet(self._legacy_fixture())

        for cell in worksheet[1]:
            self.assertTrue(cell.font.bold)
            self.assertEqual(cell.fill.fgColor.rgb, "00D9EAF7")
            self.assertEqual(cell.alignment.horizontal, "center")
            self.assertEqual(cell.alignment.vertical, "center")
            self.assertTrue(cell.alignment.wrap_text)
        for row in worksheet.iter_rows():
            for cell in row:
                self.assertEqual(cell.border.left.style, "thin")
                self.assertEqual(cell.border.left.color.rgb, "00B7B7B7")
        self.assertFalse(worksheet["F2"].alignment.wrap_text)
        for coordinate in ("G2", "H2", "I2"):
            self.assertTrue(worksheet[coordinate].alignment.wrap_text)
            self.assertEqual(worksheet[coordinate].alignment.vertical, "top")

    def test_workbook_freeze_filter_print_settings_and_margins_match_legacy(self):
        worksheet = self._worksheet(self._legacy_fixture())

        self.assertEqual(worksheet.freeze_panes, "A2")
        self.assertEqual(worksheet.auto_filter.ref, "A1:I2")
        self.assertEqual(worksheet.print_title_rows, "$1:$1")
        self.assertEqual(worksheet.page_setup.paperSize, 9)
        self.assertEqual(worksheet.page_setup.orientation, "portrait")
        self.assertEqual(worksheet.page_setup.fitToWidth, 1)
        self.assertEqual(worksheet.page_setup.fitToHeight, 0)
        self.assertTrue(worksheet.sheet_properties.pageSetUpPr.fitToPage)
        self.assertEqual(
            (
                worksheet.page_margins.left,
                worksheet.page_margins.right,
                worksheet.page_margins.top,
                worksheet.page_margins.bottom,
                worksheet.page_margins.header,
                worksheet.page_margins.footer,
            ),
            (0.2, 0.2, 0.35, 0.35, 0.15, 0.15),
        )

    def test_workbook_amount_cells_are_numeric_with_legacy_number_format(self):
        worksheet = self._worksheet(self._legacy_fixture())

        for coordinate, expected in (("D2", 215000), ("E2", 210000), ("F2", 5000)):
            self.assertEqual(worksheet[coordinate].value, expected)
            self.assertEqual(worksheet[coordinate].number_format, "#,##0")

    def test_fixed_fixture_matches_snapshot_captured_from_legacy_app(self):
        journal = self._legacy_fixture()
        expected_row = {
            "No": 1,
            "入金日": "2026/08/21",
            "取引先": "固定社",
            "入金額": 215000,
            "消込額": 210000,
            "差額": 5000,
            "未収内容": "未収運賃 100,000\n未収運賃 110,000",
            "生成仕訳": "普通預金 / 未収運賃 210,000\n普通預金 / 仮受金 5,000",
            "摘要": "固定社入金\n固定社過入金",
        }

        self.assertEqual(build_receivable_check_rows([journal]), [expected_row])
        self.assertEqual(
            build_receivable_check_filename([journal]),
            "receivable_check_20260821_abc12345.xlsx",
        )
        worksheet = self._worksheet(journal)
        self.assertEqual(
            self._workbook_snapshot(worksheet),
            {
                "title": "未収消込確認表",
                "headers": RECEIVABLE_CHECK_COLUMNS,
                "values": list(expected_row.values()),
                "widths": COLUMN_WIDTHS,
                "freeze": "A2",
                "filter": "A1:I2",
                "print_titles": "$1:$1",
                "paper_size": 9,
                "orientation": "portrait",
                "fit_width": 1,
                "fit_height": 0,
                "fit_to_page": True,
                "margins": (0.2, 0.2, 0.35, 0.35, 0.15, 0.15),
                "amount_formats": ["#,##0", "#,##0", "#,##0"],
            },
        )

    def test_service_has_no_forbidden_dependency_and_does_not_update_state_files(self):
        source = inspect.getsource(receivable_report_service)
        state_paths = [
            ROOT_DIR / "data" / "receivables" / "current.csv",
            ROOT_DIR / "data" / "receivables" / "receivable_history.csv",
            ROOT_DIR / "data" / "transactions.csv",
        ]
        before = [self._file_state(path) for path in state_paths]

        rows = build_receivable_check_rows([self._legacy_fixture()])
        build_receivable_check_filename([self._legacy_fixture()])
        build_receivable_check_excel(rows)

        after = [self._file_state(path) for path in state_paths]
        self.assertEqual(after, before)
        for forbidden in (
            "streamlit",
            "fastapi",
            "current.csv",
            "receivable_history.csv",
            "transactions.csv",
            "journal_persistence_service",
            "engine.py",
            "system_settings",
            "http",
        ):
            self.assertNotIn(forbidden, source.casefold())

    def test_streamlit_uses_service_and_keeps_existing_report_ui(self):
        app_source = (SRC_DIR / "app.py").read_text(encoding="utf-8")

        self.assertIn("from receivable_report_service import (", app_source)
        self.assertNotIn("def build_receivable_check_rows", app_source)
        self.assertNotIn("def build_receivable_check_excel", app_source)
        self.assertNotIn("def build_receivable_check_filename", app_source)
        self.assertIn('"03_未収消込確認表"', app_source)
        self.assertIn('key="save_receivable_check_excel_to_export_dir"', app_source)
        self.assertIn('key="download_receivable_check_excel"', app_source)

    @staticmethod
    def _journal(
        label,
        *,
        settlement_id=None,
        settlement_date="2026/08/21",
    ):
        return {
            "settlement_id": settlement_id or (str(label).lower() * 8)[:8],
            "settlement_date": settlement_date,
            "payment_amount": 215000,
            "target_total": 210000,
            "difference": 5000,
            "rows": [
                {
                    "借方科目": "普通預金",
                    "貸方科目": "未収運賃",
                    "金額": 210000,
                    "摘要": f"{label}社入金",
                }
            ],
            "source_candidates": [
                {
                    "未収科目": "未収運賃",
                    "消込予定": 210000,
                    "摘要": f"{label}社請求",
                }
            ],
            "customer_name": f"{label}社",
        }

    @classmethod
    def _legacy_fixture(cls):
        journal = cls._journal("固定", settlement_id="abc12345-extra")
        journal["settlement_date"] = "2026-08-21"
        journal["rows"] = [
            {
                "借方科目": "普通預金",
                "貸方科目": "未収運賃",
                "金額": 210000,
                "摘要": "固定社入金",
            },
            {
                "借方科目": "普通預金",
                "貸方科目": "仮受金",
                "金額": 5000,
                "摘要": "固定社過入金",
            },
        ]
        journal["source_candidates"] = [
            {"未収科目": "未収運賃", "消込予定": 100000, "摘要": "6月分"},
            {"未収科目": "未収運賃", "消込予定": 110000, "摘要": "7月分"},
        ]
        return journal

    @staticmethod
    def _worksheet(journal):
        rows = build_receivable_check_rows([journal])
        return load_workbook(io.BytesIO(build_receivable_check_excel(rows))).active

    @staticmethod
    def _workbook_snapshot(worksheet):
        return {
            "title": worksheet.title,
            "headers": [cell.value for cell in worksheet[1]],
            "values": [cell.value for cell in worksheet[2]],
            "widths": [
                worksheet.column_dimensions[letter].width
                for letter in "ABCDEFGHI"
            ],
            "freeze": worksheet.freeze_panes,
            "filter": worksheet.auto_filter.ref,
            "print_titles": worksheet.print_title_rows,
            "paper_size": worksheet.page_setup.paperSize,
            "orientation": worksheet.page_setup.orientation,
            "fit_width": worksheet.page_setup.fitToWidth,
            "fit_height": worksheet.page_setup.fitToHeight,
            "fit_to_page": worksheet.sheet_properties.pageSetUpPr.fitToPage,
            "margins": (
                worksheet.page_margins.left,
                worksheet.page_margins.right,
                worksheet.page_margins.top,
                worksheet.page_margins.bottom,
                worksheet.page_margins.header,
                worksheet.page_margins.footer,
            ),
            "amount_formats": [worksheet.cell(2, column).number_format for column in (4, 5, 6)],
        }

    @staticmethod
    def _file_state(path):
        if not path.exists():
            return False, None
        return True, hashlib.sha256(path.read_bytes()).hexdigest()


if __name__ == "__main__":
    unittest.main()
