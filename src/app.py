# =========================================
# 仕訳検索システム（完全版・安全版）
# 削除禁止版
# ・登録済編集維持
# ・CSV上追加
# ・エプソン45列対応
# ・AO～AS自動付与
# =========================================

import streamlit as st
import pandas as pd
import copy
import platform
import getpass
import csv
import io
import json
import os
import uuid
import re
import unicodedata

from ocr_gateway import OcrResult, PaddleOcrGateway
from ai_search.ai_search import (
    build_ai_search_context,
    run_ai_search,
)
from receivable_engine import (
    append_standard_receivables,
    apply_receivable_candidates,
    build_receivable_journal_rows,
    calculate_receivable_difference,
    convert_company_billing_excel,
    exclude_duplicate_receivables,
    is_receivable_journal_registered,
    load_receivable_history,
    load_receivables,
    mark_receivable_journal_registered,
    normalize_standard_receivable_csv,
    organize_completed_receivables,
)

DEBUG_SEARCH_DIAGNOSTICS = False

ACCOUNT_CATEGORIES = [
    "資産",
    "負債",
    "純資産",
    "収益",
    "費用",
]


def normalize_ocr_amount_value(value):

    if value in (None, ""):
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return int(value)

    normalized = unicodedata.normalize("NFKC", str(value))
    normalized = re.sub(r"[^\d]", "", normalized)

    if not normalized:
        return None

    return int(normalized)


def get_ocr_amount_candidate_value(candidate):

    if isinstance(candidate, dict):
        candidate = (
            candidate.get("amount")
            or candidate.get("value")
            or candidate.get("total")
        )

    return normalize_ocr_amount_value(candidate)


def get_representative_ocr_amount(ocr_result, amount_candidates):

    amount = normalize_ocr_amount_value(
        getattr(ocr_result, "amount", None)
    )

    if amount is not None:
        return amount

    for candidate in amount_candidates:
        amount = get_ocr_amount_candidate_value(candidate)
        if amount is not None:
            return amount

    return None


def format_ocr_amount_candidate(candidate):

    if isinstance(candidate, dict):
        amount_value = (
            candidate.get("amount")
            or candidate.get("value")
            or candidate.get("total")
        )

        parts = []
        if amount_value not in (None, ""):
            try:
                parts.append(f"¥{int(amount_value):,}")
            except Exception:
                parts.append(str(amount_value))

        for key, value in candidate.items():
            if key in {"amount", "value", "total"}:
                continue
            if value not in (None, ""):
                parts.append(f"{key}: {value}")

        return " / ".join(parts) if parts else str(candidate)

    try:
        return f"¥{int(candidate):,}"
    except Exception:
        return str(candidate)


def format_ocr_candidates(candidates, formatter=str):

    if not candidates:
        return ""

    return "\n".join(
        f"- {formatter(candidate)}"
        for candidate in candidates
    )


def is_ocr_api_connection_error(ocr_result):

    raw_text = getattr(ocr_result, "raw_text", "") or ""
    warnings = getattr(ocr_result, "warnings", None) or []

    if "OCR API接続エラー" in raw_text:
        return True

    return any(
        "OCR API接続エラー" in str(warning)
        for warning in warnings
    )


def load_account_master():

    result = {}

    for row in load_account_master_rows():

        result[
            row["name"]
        ] = row["code"]

    return result


def load_account_master_rows():

    rows = []

    with open(
        "data/account_master.csv",
        encoding="utf-8-sig"
    ) as f:

        reader = csv.DictReader(f)

        for row in reader:

            code = str(row.get("code", "")).strip()
            name = str(row.get("name", "")).strip()
            category = str(row.get("category", "")).strip()

            if code and name:
                rows.append({
                    "code": code,
                    "name": name,
                    "category": category,
                })

    return rows


def save_account_master_rows(rows):

    with open(
        "data/account_master.csv",
        "w",
        encoding="utf-8-sig",
        newline=""
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=[
                "code",
                "name",
                "category",
            ]
        )
        writer.writeheader()
        writer.writerows(rows)


ACCOUNT_MASTER = load_account_master()

def load_payment_accounts():

    result = set()

    with open(
        "data/payment_accounts.csv",
        encoding="utf-8-sig"
    ) as f:

        reader = csv.DictReader(f)

        for row in reader:

            raw_account = row.get("科目")

            if raw_account is None:
                continue

            account = str(raw_account).strip()

            if account and account.casefold() != "nan":
                result.add(account)

    return sorted(result)

def append_account_master(code, name, category=""):

    rows = load_account_master_rows()
    rows.append({
        "code": str(code).strip(),
        "name": str(name).strip(),
        "category": str(category).strip(),
    })

    save_account_master_rows(rows)

def append_payment_account(name):

    accounts = load_payment_accounts()
    name = str(name).strip()

    if name and name.casefold() != "nan":
        accounts = sorted(set(accounts + [name]))

    with open(
        "data/payment_accounts.csv",
        "w",
        encoding="utf-8",
        newline=""
    ) as f:

        writer = csv.writer(f)
        writer.writerow(["科目"])
        writer.writerows(
            [account]
            for account in accounts
        )


def add_account_candidate(
    code,
    name,
    category,
    add_to_payment_accounts=False,
):

    code = str(code).strip()
    name = str(name).strip()
    category = str(category).strip()

    if not code or not name:
        return False, [
            "科目コードと科目名を入力してください"
        ]

    if category not in ACCOUNT_CATEGORIES:
        return False, [
            "分類を選択してください"
        ]

    account_rows = load_account_master_rows()
    payment_accounts = load_payment_accounts()

    same_code = next(
        (
            row
            for row in account_rows
            if row["code"] == code
        ),
        None
    )
    same_name = next(
        (
            row
            for row in account_rows
            if row["name"] == name
        ),
        None
    )

    if same_code and same_code["name"] != name:
        return False, [
            "同じ科目コードが登録済みです"
        ]

    if same_name and same_name["code"] != code:
        return False, [
            "同じ科目名が登録済みです"
        ]

    account_registered = bool(same_code and same_name)
    messages = []
    changed = False

    if account_registered:
        messages.append(
            "科目候補には既に登録されています"
        )
    else:
        append_account_master(
            code,
            name,
            category,
        )
        messages.append(
            "科目候補に追加しました"
        )
        changed = True

    if add_to_payment_accounts:
        if name in payment_accounts:
            messages.append(
                "入金科目候補にも既に登録されています"
            )
        else:
            append_payment_account(name)
            messages.append(
                "入金科目候補に追加しました"
                if account_registered
                else "入金科目候補にも追加しました"
            )
            changed = True

    return changed, messages


def load_sub_master():

    result = {}

    with open(
        "data/sub_master.csv",
        encoding="utf-8-sig"
    ) as f:

        reader = csv.DictReader(f)

        for row in reader:

            result[
                row["name"]
            ] = row["code"]

    return result

SUB_MASTER = load_sub_master()

def load_department_master():

    result = {}

    with open(
        "data/department_master.csv",
        encoding="utf-8-sig"
    ) as f:

        reader = csv.DictReader(f)

        for row in reader:

            result[
                row["name"]
            ] = row["code"]

    return result


DEPARTMENT_MASTER = load_department_master()

from datetime import datetime

from engine import (
    EXCLUDED_SUGGESTION_ACCOUNTS,
    load_data,
    search,
    get_department,
    to_int,
    tokenize,
    get_amount_suggestions,
    get_account_suggestions,
    diagnose_debug_target,
    update_search_csv
)

from columns import (
    EPSON_COLUMNS,
    SEARCH_COLUMNS,
    EDIT_COLUMNS,
    DISPLAY_COLUMNS,
)

from columns import (
    COL_DATE,
    COL_DEBIT,
    COL_CREDIT,
    COL_DEBIT_SUB,
    COL_CREDIT_SUB,
    COL_DEBIT_AMOUNT,
    COL_CREDIT_AMOUNT,
    COL_SUMMARY,
)


# =========================================
# 科目名変換
# =========================================
def get_account_name(code):

    code = str(code)

    return ACCOUNT_MASTER.get(
        code,
        code
    )

def format_account_name_with_code(account):

    account = str(account or "").strip()

    if not account:
        return ""

    code = get_account_code(account)
    name = account

    if not code:
        for master_name, master_code in ACCOUNT_MASTER.items():
            if str(master_code).strip() == account:
                name = master_name
                code = str(master_code).strip()
                break

    return f"{name}（{code}）" if code else name

def build_account_select_options(
    records,
    summary,
    opposite_account,
    sub_account=None,
    current_account="",
    priority_accounts=None
):

    suggestions = get_account_suggestions(
        records,
        summary,
        opposite_account,
        sub_account=sub_account
    )
    recommended_accounts = [
        account
        for account, _ in suggestions
        if (
            account in account_master
            and account not in EXCLUDED_SUGGESTION_ACCOUNTS
        )
    ]
    priority_accounts = [
        account
        for account in (priority_accounts or [])
        if (
            account in account_master
            and account not in EXCLUDED_SUGGESTION_ACCOUNTS
        )
    ]
    priority_accounts = list(dict.fromkeys(priority_accounts))
    recommended_accounts = priority_accounts + [
        account
        for account in recommended_accounts
        if account not in priority_accounts
    ]

    options = recommended_accounts + [
        account
        for account in account_master
        if (
            account not in recommended_accounts
            and account not in EXCLUDED_SUGGESTION_ACCOUNTS
        )
    ]

    if (
        current_account
        and current_account not in EXCLUDED_SUGGESTION_ACCOUNTS
        and current_account not in options
    ):
        options.insert(0, current_account)

    recommended_set = set(recommended_accounts)

    return options, recommended_set


def is_excluded_account(account):

    return str(account or "").strip() in EXCLUDED_SUGGESTION_ACCOUNTS


def format_recommended_account(
    account,
    recommended_accounts,
    label="推奨"
):

    if account in recommended_accounts:
        return f"【{label}】{account}"

    return account


def extract_amount_match_detail(score_detail):

    for detail in score_detail:
        if detail.startswith("金額一致行:"):
            text = detail[len("金額一致行:"):]
            return text.rsplit(" +", 1)[0]
        if detail.startswith("金額近似:"):
            return detail.rsplit(" +", 1)[0]

    return ""


def get_matched_amount_row(rec):

    if not isinstance(rec, dict):
        return None

    matched_row = rec.get("matched_amount_row")

    if isinstance(matched_row, dict) and matched_row:
        return matched_row

    return None


def format_search_match_date(date_value):

    text = str(date_value or "").strip()

    if len(text) == 8 and text.isdigit():
        return f"{text[:4]}/{text[4:6]}/{text[6:]}"

    return text


def format_matched_row_title(matched_row):

    if not matched_row:
        return ""

    debit = matched_row.get("debit", "")
    credit = matched_row.get("credit", "")
    amount = to_int(matched_row.get("amount"))
    summary = matched_row.get("summary", "")

    parts = []

    if summary:
        parts.append(summary)

    account_text = f"{debit}→{credit}".strip("→")
    if account_text:
        parts.append(f"一致行：{account_text}")

    if amount:
        parts.append(f"¥{amount:,}")

    return " / ".join(parts)


def build_matched_row_display(matched_row):

    if not matched_row:
        return {}

    amount = to_int(matched_row.get("amount"))

    return {
        "日付": format_search_match_date(
            matched_row.get("date", "")
        ),
        "借方": format_account_name_with_code(
            matched_row.get("debit", "")
        ),
        "貸方": format_account_name_with_code(
            matched_row.get("credit", "")
        ),
        "借方補助": matched_row.get("debit_sub", ""),
        "貸方補助": matched_row.get("credit_sub", ""),
        "金額": f"{amount:,}" if amount else "",
        "摘要": matched_row.get("summary", ""),
    }


def row_has_excluded_account(row):

    if not isinstance(row, dict):
        return False

    debit = row.get(COL_DEBIT, row.get("debit", ""))
    credit = row.get(COL_CREDIT, row.get("credit", ""))

    return (
        is_excluded_account(debit)
        or is_excluded_account(credit)
    )


def get_row_account(row, side):

    if not isinstance(row, dict):
        return ""

    if side == "debit":
        return str(row.get(COL_DEBIT, row.get("debit", "")) or "").strip()

    return str(row.get(COL_CREDIT, row.get("credit", "")) or "").strip()


SETTLEMENT_REPLACEMENT_ACCOUNTS = {
    "現金",
    "普通預金",
    "当座預金",
}


def is_settlement_replacement_account(account):

    account = str(account or "").strip()

    return (
        account in SETTLEMENT_REPLACEMENT_ACCOUNTS
        or "預金" in account
        or "現金" in account
    )


def infer_block_replacement_accounts(rows, target_row, target_side):

    target_account = get_row_account(target_row, target_side)

    if not is_excluded_account(target_account):
        return []

    opposite_side = "credit" if target_side == "debit" else "debit"
    candidates = []

    for row in rows:
        same_side_account = get_row_account(row, target_side)
        opposite_account = get_row_account(row, opposite_side)

        if opposite_account != target_account:
            continue

        replacement_account = same_side_account
        if (
            replacement_account
            and not is_excluded_account(replacement_account)
            and replacement_account != target_account
        ):
            candidates.append(replacement_account)

    candidates = list(dict.fromkeys(candidates))

    return sorted(
        candidates,
        key=lambda account: (
            0
            if is_settlement_replacement_account(account)
            else 1
        )
    )


def get_account_label_set(recommended_accounts, priority_accounts):

    if priority_accounts is not None:
        return set(priority_accounts)

    return recommended_accounts


def should_show_voucher_block(rec, matched_row):

    if row_has_excluded_account(matched_row):
        return True

    for row in rec.get("rows", []):
        if row_has_excluded_account(row):
            return True

    return False


def build_voucher_block_display(rows):

    display_rows = []

    for row in rows:
        amount = max(
            to_int(row.get(COL_DEBIT_AMOUNT)),
            to_int(row.get(COL_CREDIT_AMOUNT))
        )

        display_rows.append({
            "日付": format_search_match_date(
                row.get(COL_DATE, "")
            ),
            "借方": format_account_name_with_code(
                row.get(COL_DEBIT, "")
            ),
            "貸方": format_account_name_with_code(
                row.get(COL_CREDIT, "")
            ),
            "借方補助": row.get(COL_DEBIT_SUB, ""),
            "貸方補助": row.get(COL_CREDIT_SUB, ""),
            "金額": f"{amount:,}" if amount else "",
            "摘要": row.get(COL_SUMMARY, ""),
            "伝票摘要": row.get("伝票摘要", ""),
        })

    return display_rows


RECEIVABLE_DIFFERENCE_RECOMMEND_EXCLUDED = {
    "資金複合",
    "諸口",
    "普通預金",
    "当座預金",
    "現金",
    "未収運賃",
    "未収金",
    "売掛金",
}

RECEIVABLE_OVERPAID_RECOMMEND_EXCLUDED = {
    "未払金",
    "買掛金",
    "未払費用",
    "預り金",
}


def is_receivable_difference_recommend_excluded(account):

    account = str(account or "").strip()

    if (
        account in RECEIVABLE_DIFFERENCE_RECOMMEND_EXCLUDED
        or is_excluded_account(account)
    ):
        return True

    return (
        "未収" in account
        or "売掛" in account
    )


def is_receivable_difference_recommendable(
    account,
    side,
    account_categories
):

    if is_receivable_difference_recommend_excluded(account):
        return False

    if (
        side == "credit"
        and account in RECEIVABLE_OVERPAID_RECOMMEND_EXCLUDED
    ):
        return False

    category = account_categories.get(account, "")

    if category:
        if side == "debit":
            return category == "費用"

        return category in {"負債", "収益"}

    code = str(ACCOUNT_MASTER.get(account, "")).strip()

    try:
        code_number = int(code)
    except Exception:
        code_number = 0

    if side == "debit":
        return (
            400 <= code_number < 600
            or account in {"支払手数料", "雑費"}
        )

    return (
        200 <= code_number < 300
        or account in {"仮受金", "雑収入"}
        or account.endswith("収入")
    )


def build_receivable_difference_account_options(
    records,
    customer_name,
    candidates,
    side,
    default_account,
    top_n=5
):

    allowed_accounts = [
        account
        for account in account_master
        if not is_excluded_account(account)
    ]

    if not allowed_accounts:
        return [], set(), ""

    if default_account not in allowed_accounts:
        default_account = allowed_accounts[0]

    context_text = " ".join(
        [str(customer_name or "")]
        + [
            str(candidate.get(column, "") or "")
            for candidate in candidates
            for column in [
                "未収科目",
                "未収補助",
                "部門",
                "摘要",
            ]
        ]
    )
    context_tokens = set(tokenize(context_text))

    target_column = COL_DEBIT if side == "debit" else COL_CREDIT
    account_categories = {
        row["name"]: row.get("category", "")
        for row in load_account_master_rows()
    }

    scores = {}
    counts = {}
    first_seen = {}

    for rec in records:
        for row in rec.get("rows", []):
            account = str(row.get(target_column, "") or "").strip()

            if (
                not account
                or account not in allowed_accounts
                or is_excluded_account(account)
                or not is_receivable_difference_recommendable(
                    account,
                    side,
                    account_categories
                )
            ):
                continue

            row_text = " ".join([
                str(row.get(COL_SUMMARY, "") or ""),
                str(row.get("伝票摘要", "") or ""),
                str(row.get(COL_DEBIT_SUB, "") or ""),
                str(row.get(COL_CREDIT_SUB, "") or ""),
                str(row.get(COL_DEBIT, "") or ""),
                str(row.get(COL_CREDIT, "") or ""),
            ])
            row_tokens = set(tokenize(row_text))
            score = len(context_tokens & row_tokens)

            if score <= 0:
                continue

            if account not in first_seen:
                first_seen[account] = len(first_seen)

            scores[account] = scores.get(account, 0) + score
            counts[account] = counts.get(account, 0) + 1

    recommended_accounts = sorted(
        scores,
        key=lambda account: (
            -scores[account],
            -counts[account],
            first_seen[account]
        )
    )[:top_n]

    if (
        is_receivable_difference_recommendable(
            default_account,
            side,
            account_categories
        )
        and default_account not in recommended_accounts
    ):
        recommended_accounts.insert(0, default_account)
        recommended_accounts = recommended_accounts[:top_n]

    options = recommended_accounts + [
        account
        for account in [default_account]
        if account not in recommended_accounts
    ] + [
        account
        for account in allowed_accounts
        if (
            account not in recommended_accounts
            and account != default_account
        )
    ]

    return options, set(recommended_accounts), default_account


def format_receivable_date(value):

    if hasattr(value, "strftime"):
        return value.strftime("%Y%m%d")

    return str(value or "").replace("/", "").replace("-", "")


def get_receivable_template_candidates(journal, source_candidates):

    receivable_matches = []

    for candidate in source_candidates:
        if (
            journal.get("貸方科目") == candidate.get("未収科目")
            and journal.get("貸方補助", "") == candidate.get("未収補助", "")
        ):
            receivable_matches.append(candidate)

    return receivable_matches or source_candidates


def build_empty_epson_template_row():

    return {
        column: ""
        for column in EPSON_COLUMNS
    }


def normalize_receivable_template_text(value):

    text = unicodedata.normalize("NFKC", str(value or ""))
    text = text.lower()

    for pattern in [
        "株式会社",
        "有限会社",
        "(株)",
        "（株）",
        "株)",
        "(有)",
        "（有）",
        "㈱",
        "㈲",
    ]:
        text = text.replace(pattern.lower(), "")

    text = re.sub(r"\d+\s*月分?", "", text)
    text = re.sub(r"[0-9０-９]+号車", "", text)
    text = re.sub(r"(総務課|経理課|御中|様)$", "", text)
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[　\s・･.,，．、。()\[\]（）【】\-ー－]", "", text)

    return text


def is_receivable_text_close(left, right):

    left = normalize_receivable_template_text(left)
    right = normalize_receivable_template_text(right)

    if not left or not right:
        return False

    if left in right or right in left:
        return True

    min_length = min(len(left), len(right))

    for size in range(min(4, min_length), 1, -1):
        for index in range(0, len(left) - size + 1):
            if left[index:index + size] in right:
                return True

    return False


def is_receivable_account_name(account):

    account = str(account or "")

    return (
        account in {"未収運賃", "未収金", "売掛金"}
        or "未収" in account
        or "売掛" in account
    )


def is_cash_account_name(account):

    return str(account or "") in {
        "普通預金",
        "当座預金",
        "現金",
    }


def get_receivable_template_quality(row):

    quality_columns = [
        "形式",
        "作成方法",
        "借方部門",
        "借方部門名",
        "借方消費税コード",
        "借方消費税業種",
        "借方消費税税率",
        "借方資金区分",
        "借方補助",
        "借方補助科目名",
        "貸方部門",
        "貸方部門名",
        "貸方消費税コード",
        "貸方消費税業種",
        "貸方消費税税率",
        "貸方資金区分",
        "貸方補助",
        "貸方補助科目名",
        "伝票摘要",
        COL_SUMMARY,
        "入力アプリ",
    ]

    return sum(
        1
        for column in quality_columns
        if str(row.get(column, "") or "").strip()
    )


def parse_receivable_template_date(row):

    value = str(row.get(COL_DATE, "") or "").strip()

    for date_format in ("%Y%m%d", "%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, date_format)
        except ValueError:
            continue

    return datetime.min


def is_receivable_sub_template_match(journal_sub, row_sub):

    journal_sub = str(journal_sub or "").strip()
    row_sub = str(row_sub or "").strip()

    return journal_sub == row_sub


def is_receivable_summary_template_match(journal_summary, row):

    journal_summary = str(journal_summary or "").strip()

    if not journal_summary:
        return True

    row_text = " ".join([
        str(row.get(COL_SUMMARY, "") or ""),
        str(row.get("伝票摘要", "") or ""),
    ])

    if is_receivable_text_close(journal_summary, row_text):
        return True

    journal_tokens = set(tokenize(journal_summary))
    row_tokens = set(tokenize(row_text))

    return bool(journal_tokens & row_tokens)


def is_receivable_template_text_match(search_terms, row):

    row_text = " ".join([
        str(row.get(COL_SUMMARY, "") or ""),
        str(row.get("伝票摘要", "") or ""),
        str(row.get(COL_DEBIT_SUB, "") or ""),
        str(row.get(COL_CREDIT_SUB, "") or ""),
    ])
    row_tokens = set(tokenize(row_text))

    for term in search_terms:
        term = str(term or "").strip()

        if not term:
            continue

        if is_receivable_text_close(term, row_text):
            return True

        term_tokens = set(tokenize(term))
        if term_tokens & row_tokens:
            return True

    return False


def find_receivable_template_match(
    journal,
    customer_name="",
    source_candidates=None
):

    debit_account = str(journal.get("借方科目", "") or "").strip()
    credit_account = str(journal.get("貸方科目", "") or "").strip()
    debit_sub = str(journal.get("借方補助", "") or "").strip()
    credit_sub = str(journal.get("貸方補助", "") or "").strip()
    summary = str(journal.get("摘要", "") or "").strip()
    source_candidates = source_candidates or []
    search_terms = [
        customer_name,
        summary,
    ]

    for candidate in source_candidates:
        if not isinstance(candidate, dict):
            continue

        for column in ["取引先", "得意先名", "摘要", "未収補助"]:
            value = str(candidate.get(column, "") or "").strip()
            if value:
                search_terms.append(value)

    diagnostic = {
        "DB雛形": "なし",
        "理由": "",
        "採用理由": "",
        "借方科目": debit_account,
        "貸方科目": credit_account,
        "借方補助": debit_sub,
        "貸方補助": credit_sub,
        "取引先": str(customer_name or "").strip(),
        "摘要": summary,
        "科目一致": 0,
        "借方補助一致": 0,
        "貸方補助一致": 0,
        "補助一致": "なし",
        "摘要一致": 0,
        "摘要一致有無": "なし",
        "候補数": 0,
        "採用スコア": 0,
        "雛形日付": "",
        "雛形摘要": "",
        "雛形伝票摘要": "",
        "雛形形式": "",
        "雛形借方部門": "",
        "雛形貸方部門": "",
        "雛形借方税コード": "",
        "雛形貸方税コード": "",
        "雛形品質": 0,
    }

    if not debit_account or not credit_account:
        diagnostic["理由"] = "生成仕訳の借方科目または貸方科目が空欄です"
        return None, diagnostic

    best_row = None
    best_rank = None
    best_score = 0
    best_debit_sub_matches = False
    best_credit_sub_matches = False
    best_text_matches = False

    for rec in records:
        for row in rec.get("rows", []):

            if (
                str(row.get(COL_DEBIT, "") or "").strip()
                != debit_account
                or str(row.get(COL_CREDIT, "") or "").strip()
                != credit_account
            ):
                continue

            diagnostic["科目一致"] += 1

            debit_sub_matches = (
                bool(debit_sub)
                and is_receivable_sub_template_match(
                    debit_sub,
                    row.get(COL_DEBIT_SUB, "")
                )
            )

            if debit_sub_matches:
                diagnostic["借方補助一致"] += 1

            credit_sub_matches = (
                bool(credit_sub)
                and is_receivable_sub_template_match(
                    credit_sub,
                    row.get(COL_CREDIT_SUB, "")
                )
            )

            if credit_sub_matches:
                diagnostic["貸方補助一致"] += 1

            text_matches = is_receivable_template_text_match(
                search_terms,
                row
            )

            if not text_matches:
                continue

            diagnostic["摘要一致"] += 1
            diagnostic["候補数"] += 1
            quality_score = get_receivable_template_quality(row)
            score = (
                quality_score * 10
                + (30 if text_matches else 0)
                + (10 if debit_sub_matches else 0)
                + (10 if credit_sub_matches else 0)
            )
            rank = (
                -quality_score,
                -score,
                -parse_receivable_template_date(row).toordinal(),
            )

            if best_rank is None or rank < best_rank:
                best_rank = rank
                best_row = row
                best_score = score
                best_debit_sub_matches = debit_sub_matches
                best_credit_sub_matches = credit_sub_matches
                best_text_matches = text_matches

    if best_row is None:
        if diagnostic["科目一致"] == 0:
            reason = "同じ借方科目・貸方科目の過去仕訳がありません"
        else:
            reason = "科目一致したが摘要/補助に取引先名がなく不採用"

        diagnostic["理由"] = reason
        return None, diagnostic

    diagnostic["DB雛形"] = "あり"
    diagnostic["理由"] = "生成仕訳に近い過去仕訳をDB雛形として参照します"
    diagnostic["摘要一致有無"] = "あり" if best_text_matches else "なし"
    diagnostic["補助一致"] = (
        "借方・貸方"
        if best_debit_sub_matches and best_credit_sub_matches
        else "借方"
        if best_debit_sub_matches
        else "貸方"
        if best_credit_sub_matches
        else "なし"
    )
    diagnostic["採用スコア"] = best_score
    diagnostic["雛形日付"] = best_row.get(COL_DATE, "")
    diagnostic["雛形摘要"] = best_row.get(COL_SUMMARY, "")
    diagnostic["雛形伝票摘要"] = best_row.get("伝票摘要", "")
    diagnostic["雛形形式"] = best_row.get("形式", "")
    diagnostic["雛形借方部門"] = best_row.get("借方部門名", "")
    diagnostic["雛形貸方部門"] = best_row.get("貸方部門名", "")
    diagnostic["雛形借方税コード"] = best_row.get("借方消費税コード", "")
    diagnostic["雛形貸方税コード"] = best_row.get("貸方消費税コード", "")
    diagnostic["雛形品質"] = get_receivable_template_quality(best_row)
    if diagnostic["補助一致"] == "なし":
        diagnostic["採用理由"] = "補助不一致だが摘要一致で雛形採用"
    else:
        diagnostic["採用理由"] = "科目一致・摘要一致で雛形採用"

    if diagnostic["雛形品質"] == 0:
        diagnostic["採用理由"] += "（候補はあったが雛形品質が低い）"

    return best_row, diagnostic


def find_receivable_template_row(journal, source_candidates, customer_name):

    best_row, _ = find_receivable_template_match(
        journal,
        customer_name,
        source_candidates
    )

    if best_row is None:
        return build_empty_epson_template_row()

    return {
        column: best_row.get(column, "")
        for column in EPSON_COLUMNS
    }


def build_receivable_transaction_row(
    journal,
    settlement_date,
    settlement_id,
    source_candidates,
    customer_name
):

    row = find_receivable_template_row(
        journal,
        source_candidates,
        customer_name
    )

    summary = str(journal.get("摘要", "") or "")
    debit_account = str(journal.get("借方科目", "") or "")
    credit_account = str(journal.get("貸方科目", "") or "")
    debit_sub = str(journal.get("借方補助", "") or "")
    credit_sub = str(journal.get("貸方補助", "") or "")
    department = str(journal.get("部門", "") or "")
    amount = str(journal.get("金額", "") or "")
    debit_sub_code = SUB_MASTER.get(debit_sub, "") if debit_sub else ""
    credit_sub_code = SUB_MASTER.get(credit_sub, "") if credit_sub else ""

    row[COL_DATE] = format_receivable_date(settlement_date)

    row["借方部門"] = ""
    row["借方部門名"] = ""
    row["借方科目"] = get_account_code(debit_account)
    row[COL_DEBIT] = debit_account
    row["借方補助"] = debit_sub_code
    row[COL_DEBIT_SUB] = debit_sub if debit_sub_code else ""

    row["貸方部門"] = (
        DEPARTMENT_MASTER.get(department, "")
        if department
        else ""
    )
    row["貸方部門名"] = department
    row["貸方科目"] = get_account_code(credit_account)
    row[COL_CREDIT] = credit_account
    row["貸方補助"] = credit_sub_code
    row[COL_CREDIT_SUB] = credit_sub if credit_sub_code else ""

    row[COL_DEBIT_AMOUNT] = amount
    row[COL_CREDIT_AMOUNT] = amount
    row[COL_SUMMARY] = summary

    row["証番号"] = settlement_id
    row["入力マシン"] = platform.node()
    row["入力ユーザ"] = getpass.getuser()
    row["入力アプリ"] = "仕訳検索システム"
    row["入力会社"] = st.session_state.get("company_name", "")
    row["入力日付"] = datetime.now().strftime("%Y%m%d")

    return row

# =========================================
# 伝票合計
# =========================================
def get_voucher_total(rows):

    total = 0

    for r in rows:

        try:
            total += int(
                str(r.get(COL_DEBIT_AMOUNT, 0))
                .replace(",", "")
            )
        except:
            pass

    return total

# =========================================
# 伝票分割
# =========================================
def split_journal(rows):

    debits = []
    credits = []

    for r in rows:

        d_amt = to_int(r.get(COL_DEBIT_AMOUNT))
        c_amt = to_int(r.get(COL_CREDIT_AMOUNT))

        if d_amt > 0:
            debits.append((r, d_amt))

        if c_amt > 0:
            credits.append((r, c_amt))

    # 1対多
    if len(debits) == 1 and len(credits) > 1:

        d_row, _ = debits[0]

        result = []

        for c_row, c_amt in credits:

            new = copy.deepcopy(d_row)

            new["貸方科目名"] = c_row.get(COL_CREDIT, "")
            new["貸方補助科目名"] = c_row.get(COL_CREDIT_SUB, "")
            new["貸方金額"] = str(c_amt)
            new["借方金額"] = str(c_amt)

            result.append(new)

        return result

    # 多対1
    elif len(credits) == 1 and len(debits) > 1:

        c_row, _ = credits[0]

        result = []

        for d_row, d_amt in debits:

            new = copy.deepcopy(d_row)

            new["貸方科目名"] = c_row.get(COL_CREDIT, "")
            new["貸方補助科目名"] = c_row.get(COL_CREDIT_SUB, "")
            new["貸方金額"] = str(d_amt)
            new["借方金額"] = str(d_amt)

            result.append(new)

        return result

    return rows


INPUT_CSV_COLUMNS = [
    "No",
    "伝票日付",
    "借方科目",
    "借方補助",
    "借方金額",
    "貸方科目",
    "貸方補助",
    "貸方金額",
    "摘要",
    "伝票摘要",
    "区分",
    "注意",
]


def iter_confirmed_journal_rows(confirmed_journals):

    for item in confirmed_journals:
        if isinstance(item, dict):
            yield item
        elif isinstance(item, list):
            for row in item:
                if isinstance(row, dict):
                    yield row


def build_input_csv_rows(confirmed_journals):
    """
    登録済仕訳から、手入力・確認用の簡易CSV行を作成する。
    エプソン45列取込CSVとは別物。
    """

    rows = []

    def first_value(journal, *keys):

        for key in keys:
            value = journal.get(key, "")
            if value not in (None, ""):
                return value

        return ""

    for idx, journal in enumerate(
        iter_confirmed_journal_rows(confirmed_journals),
        start=1
    ):

        debit_account = first_value(
            journal,
            COL_DEBIT,
            "借方科目名",
            "借方科目"
        )
        debit_sub = first_value(
            journal,
            COL_DEBIT_SUB,
            "借方補助科目名",
            "借方補助"
        )
        debit_amount = first_value(
            journal,
            COL_DEBIT_AMOUNT,
            "借方金額",
            "金額"
        )
        debit_sub_code = first_value(
            journal,
            "借方補助コード",
            "借方補助"
        )

        credit_account = first_value(
            journal,
            COL_CREDIT,
            "貸方科目名",
            "貸方科目"
        )
        credit_sub = first_value(
            journal,
            COL_CREDIT_SUB,
            "貸方補助科目名",
            "貸方補助"
        )
        credit_amount = first_value(
            journal,
            COL_CREDIT_AMOUNT,
            "貸方金額",
            "金額"
        )
        credit_sub_code = first_value(
            journal,
            "貸方補助コード",
            "貸方補助"
        )

        description = first_value(journal, COL_SUMMARY, "摘要")
        voucher_description = first_value(journal, "伝票摘要")

        note_items = []

        if (
            journal.get("DB雛形") == "なし"
            or journal.get("db_template_found") is False
        ):
            note_items.append("DB雛形なし")

        if debit_sub and not debit_sub_code:
            note_items.append("借方補助コード未取得")

        if credit_sub and not credit_sub_code:
            note_items.append("貸方補助コード未取得")

        if not voucher_description:
            note_items.append("伝票摘要なし")

        rows.append({
            "No": idx,
            "伝票日付": first_value(journal, COL_DATE, "伝票日付", "日付"),
            "借方科目": debit_account,
            "借方補助": debit_sub,
            "借方金額": debit_amount,
            "貸方科目": credit_account,
            "貸方補助": credit_sub,
            "貸方金額": credit_amount,
            "摘要": description,
            "伝票摘要": voucher_description,
            "区分": first_value(journal, "区分", "source", "処理区分"),
            "注意": " / ".join(note_items),
        })

    return rows


def build_ai_search_candidates(results):

    candidates = []
    score_details = []

    for index, result in enumerate(results or [], start=1):
        if len(result) != 3:
            continue

        score, rec, score_detail = result

        if not isinstance(rec, dict):
            continue

        rows = rec.get("rows", [])
        if not rows:
            continue

        first_row = rows[0]
        candidate = {
            "rank": index,
            "score": score,
            "date": first_row.get(COL_DATE, ""),
            "summary": first_row.get(COL_SUMMARY, ""),
            "debit": first_row.get(COL_DEBIT, ""),
            "credit": first_row.get(COL_CREDIT, ""),
            "row_count": len(rows),
            "amount": get_voucher_total(rows),
        }
        candidates.append(candidate)

        if score_detail:
            score_details.append({
                "rank": index,
                "detail": list(score_detail),
            })

    return candidates, score_details

# =========================================
# エプソンCSV変換
# =========================================
def build_epson_rows(rows, company_name):

    result = []

    machine_name = platform.node()
    user_name = getpass.getuser()

    app_name = "仕訳検索システム"

    input_date = datetime.now().strftime("%Y%m%d")

    for r in rows:

        row = {
            c: r.get(c, "")
            for c in EPSON_COLUMNS
        }

        summary = r.get(COL_SUMMARY, "")
        debit_sub_name = r.get(COL_DEBIT_SUB, "")
        credit_sub_name = r.get(COL_CREDIT_SUB, "")

        # =====================================
        # 画面で編集した項目だけ上書き
        # =====================================
        row["伝票日付"] = r.get(COL_DATE, "")
        row["摘要"] = summary

        # 伝票摘要はDB雛形の値を保持する。
        # 摘要から伝票摘要への自動コピーは禁止。

        # =====================================
        # 借方
        # =====================================
        row["借方科目"] = get_account_code(
            r.get(COL_DEBIT, "")
        ) or r.get("借方科目", "")
        row["借方科目名"] = r.get(COL_DEBIT, "")

        row["借方補助"] = (
            SUB_MASTER.get(debit_sub_name, r.get("借方補助", ""))
            if debit_sub_name
            else ""
        )
        row["借方補助科目名"] = debit_sub_name

        row["借方金額"] = r.get(COL_DEBIT_AMOUNT, "")

        # =====================================
        # 貸方
        # =====================================
        row["貸方科目"] = get_account_code(
            r.get(COL_CREDIT, "")
        ) or r.get("貸方科目", "")
        row["貸方科目名"] = r.get(COL_CREDIT, "")

        row["貸方補助"] = (
            SUB_MASTER.get(credit_sub_name, r.get("貸方補助", ""))
            if credit_sub_name
            else ""
        )
        row["貸方補助科目名"] = credit_sub_name

        row["貸方金額"] = r.get(COL_CREDIT_AMOUNT, "")

        # =====================================
        # AO～AS
        # =====================================
        row["入力マシン"] = machine_name
        row["入力ユーザ"] = user_name
        row["入力アプリ"] = app_name
        row["入力会社"] = company_name
        row["入力日付"] = input_date

        result.append(row)

    return result


def save_exported_journals(rows):

    duplicate_columns = [
        COL_DATE,
        "借方科目",
        COL_DEBIT,
        COL_DEBIT_SUB,
        "貸方科目",
        COL_CREDIT,
        COL_CREDIT_SUB,
        COL_DEBIT_AMOUNT,
        COL_SUMMARY,
    ]

    def row_key(row):

        values = []

        for column in duplicate_columns:
            value = " ".join(
                str(row.get(column, "")).split()
            )

            if column == COL_DATE:
                value = value.replace("/", "").replace("-", "")
            elif column == COL_DEBIT_AMOUNT:
                value = value.replace(",", "")

            values.append(value)

        return tuple(values)

    try:
        existing_df = pd.read_csv(
            "data/transactions.csv",
            dtype=str
        ).fillna("")
    except (FileNotFoundError, pd.errors.EmptyDataError):
        existing_df = pd.DataFrame()

    registered_keys = {
        row_key(row)
        for _, row in existing_df.iterrows()
    }
    new_rows = []

    for row in rows:
        key = row_key(row)

        if key in registered_keys:
            continue

        registered_keys.add(key)
        new_rows.append(row)

    if new_rows:
        update_search_csv([new_rows])

    st.session_state["epson_export_success"] = (
        "エプソンCSVを作成し、検索DBも更新しました"
    )
    st.cache_data.clear()


TRANSACTIONS_PATH = "data/transactions.csv"


def read_past_journal_csv(content):

    last_error = None

    for encoding in [
        "utf-8-sig",
        "cp932",
    ]:
        try:
            df = pd.read_csv(
                io.BytesIO(content),
                dtype=str,
                encoding=encoding,
            ).fillna("")
            df.columns = [
                str(column).strip()
                for column in df.columns
            ]
            return df, encoding, None
        except Exception as e:
            last_error = e

    return None, None, last_error


def load_transactions_df():

    return pd.read_csv(
        TRANSACTIONS_PATH,
        dtype=str,
        encoding="utf-8-sig",
    ).fillna("")


def normalize_import_value(value, amount=False, date=False):

    value = " ".join(
        str(value or "").split()
    )

    if date:
        value = value.replace("/", "").replace("-", "")

    if amount:
        value = value.replace(",", "")

    return value


def journal_import_key(row):

    amount = row.get(COL_DEBIT_AMOUNT, "")

    if not normalize_import_value(amount, amount=True):
        amount = row.get(COL_CREDIT_AMOUNT, "")

    return (
        normalize_import_value(row.get(COL_DATE, ""), date=True),
        normalize_import_value(row.get("借方科目", "")),
        normalize_import_value(row.get(COL_DEBIT, "")),
        normalize_import_value(row.get("貸方科目", "")),
        normalize_import_value(row.get(COL_CREDIT, "")),
        normalize_import_value(amount, amount=True),
        normalize_import_value(row.get(COL_SUMMARY, "")),
    )


def prepare_past_journal_import(upload_df, existing_df):

    existing_columns = list(existing_df.columns)

    if len(upload_df.columns) != 45:
        return None, f"45列CSVではありません（{len(upload_df.columns)}列）"

    if list(upload_df.columns) != existing_columns:
        return None, "CSVの列構造がtransactions.csvと一致しません"

    upload_df = upload_df[
        upload_df.apply(
            lambda row: any(
                str(value).strip()
                for value in row
            ),
            axis=1,
        )
    ]

    existing_keys = {
        journal_import_key(row)
        for _, row in existing_df.iterrows()
    }

    seen_import_keys = set()
    new_rows = []
    duplicate_count = 0

    for _, row in upload_df.iterrows():
        key = journal_import_key(row)

        if key in existing_keys or key in seen_import_keys:
            duplicate_count += 1
            continue

        seen_import_keys.add(key)
        new_rows.append(row.to_dict())

    new_df = pd.DataFrame(
        new_rows,
        columns=existing_columns,
    )

    return {
        "new_df": new_df,
        "duplicate_count": duplicate_count,
    }, None


def append_past_journals_to_transactions(new_df):

    if new_df.empty:
        return 0

    existing_df = load_transactions_df()
    combined_df = pd.concat(
        [
            existing_df,
            new_df,
        ],
        ignore_index=True,
    )

    combined_df.to_csv(
        TRANSACTIONS_PATH,
        index=False,
        encoding="utf-8-sig",
    )

    st.cache_data.clear()

    return len(new_df)


def save_csv_to_export_dir(csv_bytes, filename, export_dir):

    export_dir = str(export_dir or "").strip()

    if not export_dir:
        return False, "CSV保存先フォルダを入力してください"

    if not os.path.isdir(export_dir):
        return False, "CSV保存先フォルダが存在しません"

    save_path = os.path.join(
        export_dir,
        os.path.basename(filename)
    )

    try:
        with open(save_path, "wb") as file:
            file.write(csv_bytes)
    except OSError as e:
        return False, f"CSVを保存できませんでした: {e}"

    return True, f"保存しました：{save_path}"


def save_file_to_export_dir(file_bytes, filename, export_dir):

    export_dir = str(export_dir or "").strip()

    if not export_dir:
        return False, "CSV保存先フォルダを入力してください"

    if not os.path.isdir(export_dir):
        return False, "CSV保存先フォルダが存在しません"

    save_path = os.path.join(
        export_dir,
        os.path.basename(filename)
    )

    try:
        with open(save_path, "wb") as file:
            file.write(file_bytes)
    except OSError as e:
        return False, f"ファイルを保存できませんでした: {e}"

    return True, f"保存しました：{save_path}"


def build_input_journal_excel(input_df):

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.properties import PageSetupProperties

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "簡易仕訳帳"

    headers = list(input_df.columns)
    worksheet.append(headers)

    amount_columns = {
        "借方金額",
        "貸方金額",
    }
    wrap_columns = {
        "摘要",
        "伝票摘要",
        "注意",
    }
    wrap_column_letters = set()

    for _, row in input_df.iterrows():
        values = []
        for column in headers:
            value = row.get(column, "")
            if column in amount_columns and value not in ("", None):
                try:
                    value = int(str(value).replace(",", ""))
                except ValueError:
                    pass
            values.append(value)
        worksheet.append(values)

    column_widths = {
        "No": 4,
        "伝票日付": 10,
        "借方科目": 12,
        "借方補助": 14,
        "借方金額": 11,
        "貸方科目": 12,
        "貸方補助": 14,
        "貸方金額": 11,
        "摘要": 24,
        "伝票摘要": 18,
        "区分": 8,
        "注意": 18,
    }

    thin_side = Side(style="thin", color="B7B7B7")
    cell_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7"
    )

    for column_index, column_name in enumerate(headers, start=1):
        column_letter = worksheet.cell(
            row=1,
            column=column_index
        ).column_letter
        worksheet.column_dimensions[column_letter].width = (
            column_widths.get(column_name, 12)
        )
        if column_name in wrap_columns:
            wrap_column_letters.add(column_letter)

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = cell_border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(cell.column_letter in wrap_column_letters)
            )
            if headers[cell.column - 1] in amount_columns and cell.row > 1:
                cell.number_format = "#,##0"

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"
    worksheet.print_title_rows = "1:1"
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    if worksheet.sheet_properties.pageSetUpPr is None:
        worksheet.sheet_properties.pageSetUpPr = PageSetupProperties()
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.5
    worksheet.page_margins.bottom = 0.5
    worksheet.page_margins.header = 0.2
    worksheet.page_margins.footer = 0.2

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


SETTINGS_PATH = os.path.join(
    "config",
    "settings.json"
)


def load_system_settings():

    try:
        with open(SETTINGS_PATH, "r", encoding="utf-8") as file:
            settings = json.load(file)
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return {}

    if not isinstance(settings, dict):
        return {}

    return {
        "company_name": str(settings.get("company_name", "") or ""),
        "csv_export_dir": str(settings.get("csv_export_dir", "") or ""),
    }


def save_system_settings(company_name, csv_export_dir):

    settings = {
        "company_name": str(company_name or ""),
        "csv_export_dir": str(csv_export_dir or ""),
    }

    try:
        os.makedirs(
            os.path.dirname(SETTINGS_PATH),
            exist_ok=True
        )
        with open(SETTINGS_PATH, "w", encoding="utf-8") as file:
            json.dump(
                settings,
                file,
                ensure_ascii=False,
                indent=2
            )
    except OSError as e:
        return False, f"システム設定を保存できませんでした: {e}"

    return True, ""


def save_system_settings_from_state():

    saved, message = save_system_settings(
        st.session_state.get("company_name", ""),
        st.session_state.get("csv_export_dir", "")
    )

    if not saved:
        st.session_state["system_settings_warning"] = message

# =========================================
# 初期設定
# =========================================
st.set_page_config(
    page_title="仕訳検索",
    layout="wide"
)

# =========================================
# データロード
# =========================================
@st.cache_data
def cached_load():
    return load_data()

records, name_to_code, freq = cached_load()

def get_account_code(account_name):

    account_name = str(account_name).strip()

    return ACCOUNT_MASTER.get(
        account_name,
        name_to_code.get(account_name, "")
    )

def keep_receivable_customer_open(customer_name):

    st.session_state[
        "open_receivable_customer"
    ] = customer_name
    st.session_state[
        "active_receivable_customer"
    ] = customer_name

def parse_receivable_payment_amount(value):

    if value is None:
        return None

    normalized = str(value)
    for old, new in [
        (",", ""),
        ("，", ""),
        ("円", ""),
        ("￥", ""),
        ("¥", "")
    ]:
        normalized = normalized.replace(old, new)

    normalized = "".join(normalized.split())

    if not normalized or not normalized.isdigit():
        return None

    amount = int(normalized)

    if amount <= 0:
        return None

    return amount

account_master = sorted(
    ACCOUNT_MASTER.keys()
)

department_master = sorted(
    DEPARTMENT_MASTER.keys()
)

sub_master = sorted(
    SUB_MASTER.keys()
)


def build_sub_options(current_sub):

    options = [""] + sub_master

    if current_sub and current_sub not in options:
        options.append(current_sub)

    return options


# =========================================
# 科目マスター生成
# =========================================
def generate_account_master(records):

    rows = []

    seen = set()

    for rec in records:

        for r in rec["rows"]:

            debit_code = str(
                r.get("借方科目", "")
            ).strip()

            debit_name = str(
                r.get(COL_DEBIT, "")
            ).strip()

            if debit_code and debit_name:

                key = (
                    debit_code,
                    debit_name
                )

                if key not in seen:

                    seen.add(key)

                    rows.append({
                        "code": debit_code,
                        "name": debit_name,
                        "category": ""
                    })

            credit_code = str(
                r.get("貸方科目", "")
            ).strip()

            credit_name = str(
                r.get(COL_CREDIT, "")
            ).strip()

            if credit_code and credit_name:

                key = (
                    credit_code,
                    credit_name
                )

                if key not in seen:

                    seen.add(key)

                    rows.append({
                        "code": credit_code,
                        "name": credit_name,
                        "category": ""
                    })

    rows = sorted(
        rows,
        key=lambda x: x["code"]
    )

    pd.DataFrame(rows).to_csv(
        "data/account_master.csv",
        index=False,
        encoding="utf-8-sig"
    )

    return len(rows)

# =========================================
# 部門マスター生成
# =========================================
def generate_department_master(records):

    rows = []
    seen = set()

    for rec in records:

        for r in rec["rows"]:

            targets = [

                (
                    str(r.get("借方部門", "")).strip(),
                    str(r.get("借方部門名", "")).strip()
                ),

                (
                    str(r.get("貸方部門", "")).strip(),
                    str(r.get("貸方部門名", "")).strip()
                )

            ]

            for code, name in targets:

                if code and name:

                    key = (code, name)

                    if key not in seen:

                        seen.add(key)

                        rows.append({
                            "code": code,
                            "name": name
                        })


    pd.DataFrame(rows).to_csv(
        "data/department_master.csv",
        index=False,
        encoding="utf-8-sig"
    )

    return len(rows)

# =========================================
# 補助マスター生成
# =========================================
def generate_sub_master(records):

    rows = []
    seen = set()

    for rec in records:

        for r in rec["rows"]:

            targets = [

                (
                    str(r.get("借方補助", "")).strip(),
                    str(r.get("借方補助科目名", "")).strip()
                ),

                (
                    str(r.get("貸方補助", "")).strip(),
                    str(r.get("貸方補助科目名", "")).strip()
                )

            ]

            for code, name in targets:

                if code and name:

                    key = (code, name)

                    if key not in seen:

                        seen.add(key)

                        rows.append({
                            "code": code,
                            "name": name
                        })

    pd.DataFrame(rows).to_csv(
        "data/sub_master.csv",
        index=False,
        encoding="utf-8-sig"
    )

    return len(rows)

mode = st.sidebar.radio(
    "モード",
    [
        "通常仕訳",
        "未収消込"
    ]
)

if mode == "通常仕訳":

    st.title("📘 仕訳検索システム")

    # =========================================
    # セッション
    # =========================================
    if "results" not in st.session_state:
        st.session_state.results = []

    if "journal_result_limit" not in st.session_state:
        st.session_state["journal_result_limit"] = 5
    
    if "confirmed" not in st.session_state:
        st.session_state.confirmed = []

    loaded_system_settings = load_system_settings()

    if "company_name" not in st.session_state:
        st.session_state.company_name = loaded_system_settings.get(
            "company_name",
            ""
        )

    if "csv_export_dir" not in st.session_state:
        st.session_state["csv_export_dir"] = loaded_system_settings.get(
            "csv_export_dir",
            ""
        )

    def reset_for_next_journal_search():

        st.session_state["keyword_input"] = ""
        st.session_state["search_amount"] = None
        st.session_state.results = []
        st.session_state.pop("ocr_search_text_pending", None)
        st.session_state.pop("ocr_search_amount_pending", None)
        st.session_state.pop("selected_candidate_index", None)
        st.session_state["selected_candidate_no"] = 1
        st.session_state["candidate_number_select_input"] = 1
        st.session_state["last_journal_search_limit"] = (
            st.session_state.get("journal_result_limit", 5)
        )

        if "editing_candidate_values" in st.session_state:
            st.session_state["editing_candidate_values"] = {}

        for key in list(st.session_state.keys()):
            if str(key).startswith("next_search_ready_"):
                st.session_state.pop(key, None)
    
    
    # =========================================
    # サイドバー
    # =========================================
    
    # -----------------------------------------
    # システム設定
    # -----------------------------------------
    st.sidebar.header("🏢 システム設定")
    
    st.sidebar.text_input(
        "入力会社",
        key="company_name",
        on_change=save_system_settings_from_state
    )

    st.sidebar.text_input(
        "CSV保存先フォルダ",
        placeholder=r"\\NAS\share\journal-ai\csv",
        key="csv_export_dir",
        on_change=save_system_settings_from_state
    )

    if "system_settings_warning" in st.session_state:
        st.sidebar.warning(
            st.session_state.pop("system_settings_warning")
        )

    if "account_candidate_success" in st.session_state:
        for message in st.session_state.pop(
            "account_candidate_success"
        ):
            st.sidebar.success(message)

    st.sidebar.divider()
    
    # -----------------------------------------
    # 検索
    # -----------------------------------------
    st.sidebar.header("🔍 検索")
    
    if "keyword_input" not in st.session_state:
        st.session_state["keyword_input"] = ""
    
    if "ocr_search_text_pending" in st.session_state:
        st.session_state["keyword_input"] = (
            st.session_state.pop("ocr_search_text_pending")
        )

    if "ocr_search_amount_pending" in st.session_state:
        st.session_state["search_amount"] = (
            st.session_state.pop("ocr_search_amount_pending")
        )
    
    with st.sidebar.form(key="journal_search_form"):
        dept = st.selectbox(
            "部門",
            [""] + department_master,
            key="search_department"
        )

        keyword = st.text_input(
            "キーワード",
            key="keyword_input"
        )

        amount_input = st.number_input(
            "金額",
            value=None,
            min_value=0,
            step=1,
            placeholder="0",
            key="search_amount"
        )

        st.caption("条件を入力し、Enterキーでも検索できます。")
        search_clicked = st.form_submit_button(
            "検索",
            type="primary"
        )

    amount = (
        int(amount_input)
        if amount_input is not None and amount_input > 0
        else None
    )
    
    if st.sidebar.button(
        "科目マスター生成"
    ):
        count = generate_account_master(
            records
        )
    
        ACCOUNT_MASTER.clear()
    
        ACCOUNT_MASTER.update(
            load_account_master()
        )
    
        st.toast(
            f"科目マスター {count}件生成しました"
        )
    
    
    if st.sidebar.button(
        "部門マスター生成"
    ):
        
        count = generate_department_master(
            records
        )
    
        DEPARTMENT_MASTER.clear()
    
        DEPARTMENT_MASTER.update(
            load_department_master()
        )
    
        st.toast(
            f"部門マスター {count}件生成しました"
        )
    
    
    if st.sidebar.button(
        "補助マスター生成"
    ):
    
        count = generate_sub_master(
            records
        )
    
        SUB_MASTER.clear()
    
        SUB_MASTER.update(
            load_sub_master()
        )
    
        st.toast(
            f"補助マスター {count}件生成しました"
        )
    
    # =========================================
    # 検索実行
    # =========================================
    if search_clicked:

        result_limit = int(
            st.session_state.get("journal_result_limit", 5)
        )
        search_params = {
            "keyword": st.session_state["keyword_input"],
            "dept": dept if dept else None,
            "amount": amount,
        }
    
        st.session_state.results = search(
            records,
            search_params["keyword"],
            search_params["dept"],
            search_params["amount"],
            freq,
            limit=result_limit
        )
        st.session_state["last_journal_search_params"] = search_params
        st.session_state["last_journal_search_limit"] = result_limit
        st.session_state.pop("selected_candidate_index", None)
    
    (
        date_label_col,
        date_input_col,
        ocr_label_col,
        ocr_upload_col,
    ) = st.columns([0.9, 1.6, 0.9, 3.2])

    # =========================================
    # 日付
    # =========================================
    with date_label_col:
        st.markdown("**伝票日付**")

    with date_input_col:
        process_date_obj = st.date_input(
            "日付",
            datetime.today(),
            label_visibility="collapsed"
        )

    process_date = process_date_obj.strftime("%Y%m%d")

    with st.sidebar.expander("科目マスター管理"):

        st.subheader("科目候補の追加")

        st.caption(
            "※ エプソン側に登録済みの科目だけを追加してください。"
        )
        st.caption(
            "※ journal-aiで追加しても、エプソン側には登録されません。"
        )
        st.caption(
            "※ 分類は検索・確認・AIチェック用の補助情報です。"
        )

        candidate_code = st.text_input(
            "科目コード",
            key="sidebar_account_candidate_code"
        ).strip()

        candidate_name = st.text_input(
            "科目名",
            key="sidebar_account_candidate_name"
        ).strip()

        candidate_category = st.selectbox(
            "分類",
            ACCOUNT_CATEGORIES,
            key="sidebar_account_candidate_category"
        )

        candidate_add_to_payment = st.checkbox(
            "入金科目候補にも追加する",
            key="sidebar_account_candidate_payment"
        )

        if st.button(
            "追加する",
            key="sidebar_add_account_candidate"
        ):
            changed, messages = add_account_candidate(
                candidate_code,
                candidate_name,
                candidate_category,
                candidate_add_to_payment,
            )

            if changed:
                st.session_state[
                    "account_candidate_success"
                ] = messages
                st.rerun()
            else:
                for message in messages:
                    st.warning(message)

    # =========================================
    # 過去仕訳CSV取込
    # =========================================
    with st.sidebar.expander("過去仕訳CSV取込"):

        st.caption("初回設定または過去DB更新時に使用します。")

        if "past_journal_import_success" in st.session_state:
            st.success(
                st.session_state.pop(
                    "past_journal_import_success"
                )
            )

        past_journal_file = st.file_uploader(
            "エプソン仕訳CSVアップロード",
            type=["csv"],
            key="past_journal_csv_upload"
        )

        if past_journal_file is not None:
            upload_df, encoding, read_error = read_past_journal_csv(
                past_journal_file.getvalue()
            )

            if read_error is not None:
                st.error(
                    f"CSVを読み込めませんでした: {read_error}"
                )
            else:
                st.caption(
                    f"文字コード: {encoding} / {len(upload_df.columns)}列"
                )

                existing_df = load_transactions_df()
                import_result, import_error = prepare_past_journal_import(
                    upload_df,
                    existing_df,
                )

                if import_error:
                    st.warning(import_error)
                else:
                    new_df = import_result["new_df"]
                    duplicate_count = import_result["duplicate_count"]

                    st.write(
                        "新規件数:",
                        len(new_df)
                    )
                    st.write(
                        "重複件数:",
                        duplicate_count
                    )

                    preview_columns = [
                        column
                        for column in [
                            COL_DATE,
                            "借方科目",
                            COL_DEBIT,
                            "貸方科目",
                            COL_CREDIT,
                            COL_DEBIT_AMOUNT,
                            COL_SUMMARY,
                        ]
                        if column in upload_df.columns
                    ]

                    st.dataframe(
                        upload_df[preview_columns].head(50),
                        use_container_width=True
                    )

                    if st.button(
                        "検索DBへ追加",
                        disabled=new_df.empty,
                        key="append_past_journals_to_search_db"
                    ):
                        appended_count = append_past_journals_to_transactions(
                            new_df
                        )
                        st.session_state[
                            "past_journal_import_success"
                        ] = (
                            f"検索DBへ{appended_count}件追加しました"
                        )
                        st.rerun()
    
    # =========================================
    # OCR読込
    # =========================================
    with ocr_label_col:
        st.markdown("**OCR読込**")

    with ocr_upload_col:
        uploaded_file = st.file_uploader(
            "JPG / PNG / PDF",
            type=["jpg", "jpeg", "png", "pdf"],
            label_visibility="collapsed"
        )
    
    current_file = None
    
    if uploaded_file:
        current_file = uploaded_file.name
    
    if "last_uploaded_file" not in st.session_state:
        st.session_state["last_uploaded_file"] = ""
    
    if "ocr_result" not in st.session_state:
        st.session_state["ocr_result"] = None
    
    if current_file != st.session_state["last_uploaded_file"]:
    
        st.session_state["ocr_done"] = False
        st.session_state["ocr_result"] = None
        st.session_state["last_uploaded_file"] = current_file
    
    # =========================================
    # OCR実行
    # =========================================
    if uploaded_file and not st.session_state.get("ocr_done", False):
    
        st.success(
            f"アップロード: {uploaded_file.name}"
        )
    
        st.info("OCR解析中...")
    
        gateway = PaddleOcrGateway()
    
        try:
            st.session_state["ocr_result"] = gateway.analyze(
                content=uploaded_file.getvalue(),
                filename=uploaded_file.name,
                mime_type=uploaded_file.type or "",
            )
        except Exception as e:
            message = f"OCR API接続エラー: {type(e).__name__}: {e}"
            st.session_state["ocr_result"] = OcrResult(
                raw_text=message,
                confidence=0.0,
                warnings=[message],
            )
    
        # OCR済み
        st.session_state["ocr_done"] = True

        if is_ocr_api_connection_error(st.session_state["ocr_result"]):
            st.warning("OCR解析失敗：AIサーバーに接続できません")
        else:
            st.success("OCR解析完了")
    
    ocr_result = st.session_state["ocr_result"]
    
    if ocr_result is not None:
    
        search_text_candidates = (
            getattr(ocr_result, "search_text_candidates", None)
            or ([ocr_result.search_text] if ocr_result.search_text else [])
        )
        amount_candidates = (
            getattr(ocr_result, "amount_candidates", None)
            or ([ocr_result.amount] if ocr_result.amount is not None else [])
        )
        memo_candidates = (
            getattr(ocr_result, "memo_candidates", None)
            or ([ocr_result.memo] if ocr_result.memo else [])
        )
        invoice_number_candidates = (
            getattr(ocr_result, "invoice_number_candidates", None)
            or (
                [ocr_result.invoice_registration_number]
                if ocr_result.invoice_registration_number
                else []
            )
        )
        warnings = getattr(ocr_result, "warnings", None) or []
        raw_text = getattr(ocr_result, "raw_text", "") or ""
        representative_amount = get_representative_ocr_amount(
            ocr_result,
            amount_candidates,
        )

        with st.expander("OCRレスポンスデバッグ（一時）", expanded=True):
            st.write("ocr_result.amount:", getattr(ocr_result, "amount", None))
            st.write(
                "ocr_result.amount_candidates:",
                getattr(ocr_result, "amount_candidates", None),
            )
            st.write(
                "ocr_result.invoice_registration_number:",
                getattr(ocr_result, "invoice_registration_number", None),
            )
            st.write(
                "ocr_result.invoice_number_candidates:",
                getattr(ocr_result, "invoice_number_candidates", None),
            )
            st.write("ocr_result.raw_text 先頭300文字:", raw_text[:300])
            if warnings:
                st.write("warnings:", warnings)

        st.write(
            "検索文字列候補:",
            format_ocr_candidates(search_text_candidates)
        )
    
        st.write(
            "代表金額候補:",
            (
                format_ocr_amount_candidate(representative_amount)
                if representative_amount is not None
                else ""
            )
        )

        st.write(
            "金額候補一覧:",
            format_ocr_candidates(
                amount_candidates,
                format_ocr_amount_candidate,
            )
        )
    
        st.write(
            "摘要候補:",
            format_ocr_candidates(memo_candidates)
        )

        st.write(
            "T番号候補:",
            format_ocr_candidates(invoice_number_candidates)
        )

        if warnings:
            for warning in warnings:
                st.warning(str(warning))
        else:
            st.write("警告:", "")
    
        st.write(
            "OCR全文:",
            ocr_result.raw_text
        )
    
        st.write(
            "信頼度:",
            ocr_result.confidence
        )

        if ocr_result.invoice_number_present:
            invoice_registration_number = (
                ocr_result.invoice_registration_number
                or ""
            )
            invoice_format_status = (
                "OK"
                if ocr_result.invoice_number_valid_format
                else "要確認"
            )
            invoice_verified_status = (
                "確認済み"
                if ocr_result.invoice_number_verified
                else "未確認"
            )
            st.write(
                "適格請求書番号:",
                invoice_registration_number
            )

            st.write(
                "形式:",
                invoice_format_status
            )

            st.write(
                "国税庁照会:",
                invoice_verified_status
            )

            if len(invoice_number_candidates) > 1:
                st.write(
                    "候補:",
                    "、".join(invoice_number_candidates)
                )
            elif invoice_number_candidates:
                st.write(
                    "候補:",
                    invoice_number_candidates[0]
                )
    
        if st.button(
            "検索へ反映",
            disabled=not bool(ocr_result.search_text.strip()),
        ):
            st.session_state["ocr_search_text_pending"] = (
                ocr_result.search_text
            )
            st.session_state["ocr_search_amount_pending"] = (
                representative_amount
            )
            st.rerun()
    
    st.divider()
    # =========================================
    # 検索結果
    # =========================================
    results = st.session_state.results
    
    if not results:
    
        st.info("検索結果がありません")
    
    else:

        if DEBUG_SEARCH_DIAGNOSTICS:
            debug_search_params = st.session_state.get(
                "last_journal_search_params",
                {
                    "keyword": st.session_state.get("keyword_input", ""),
                    "dept": dept if dept else None,
                    "amount": amount,
                }
            )
            debug_diagnostic = diagnose_debug_target(
                records,
                debug_search_params.get("keyword", ""),
                debug_search_params.get("dept"),
                debug_search_params.get("amount"),
                freq,
                results=results
            )

            with st.expander("検索DB診断（開発用）", expanded=False):
                st.dataframe(
                    pd.DataFrame([{
                        "対象行": (
                            "あり"
                            if debug_diagnostic.get("target_exists")
                            else "なし"
                        ),
                        "rows": (
                            "あり"
                            if debug_diagnostic.get("in_rows")
                            else "なし"
                        ),
                        "search_rows": (
                            "あり"
                            if debug_diagnostic.get("in_search_rows")
                            else "なし"
                        ),
                        "スコア": debug_diagnostic.get("score"),
                        "順位": debug_diagnostic.get("rank"),
                        "所属グループID": (
                            debug_diagnostic.get("group_id")
                        ),
                        "matched_row保持": (
                            "あり"
                            if debug_diagnostic.get("matched_row_kept")
                            else "なし"
                        ),
                        "候補表示へ渡る": (
                            "あり"
                            if debug_diagnostic.get("passed_to_display")
                            else "なし"
                        ),
                    }]),
                    hide_index=True,
                    use_container_width=True
                )

                target_row = debug_diagnostic.get("target_row")
                if target_row:
                    st.markdown("**対象行**")
                    st.dataframe(
                        pd.DataFrame([target_row]),
                        hide_index=True,
                        use_container_width=True
                    )

                representative_row = debug_diagnostic.get(
                    "representative_row"
                )
                if representative_row:
                    st.markdown("**グループ代表行**")
                    st.dataframe(
                        pd.DataFrame([representative_row]),
                        hide_index=True,
                        use_container_width=True
                    )

                score_detail = debug_diagnostic.get("score_detail") or []
                if score_detail:
                    st.markdown("**スコア内訳**")
                    st.write(score_detail)

        with st.expander("AIサーチ（補足説明）", expanded=False):
            st.caption(
                "検索結果をもとに、候補選択の理由や注意点を整理します。"
                "仕訳の決定や検索順位の変更は行いません。"
            )

            if st.button(
                "AIサーチで説明を表示",
                key="run_ai_search_explanation"
            ):
                ai_candidates, ai_score_detail = build_ai_search_candidates(
                    results
                )
                context = build_ai_search_context(
                    keyword=st.session_state.get("keyword_input", ""),
                    amount=amount,
                    department=dept,
                    candidates=ai_candidates,
                    score_detail=ai_score_detail,
                    ocr_text=(
                        getattr(ocr_result, "raw_text", "")
                        if ocr_result is not None
                        else ""
                    ),
                )
                ai_result = run_ai_search(context)

                st.info(ai_result.get("summary", ""))

                reasons = ai_result.get("reason", [])
                if reasons:
                    st.markdown("**理由**")
                    for reason in reasons:
                        st.write(f"- {reason}")

                warnings = ai_result.get("warning", [])
                if warnings:
                    st.markdown("**注意点**")
                    for warning in warnings:
                        st.warning(warning)

        result_count_message = st.empty()

        (
            label_col,
            input_col,
            button_col,
            limit_label_col,
            limit_col,
            spacer_col,
        ) = st.columns([0.8, 0.65, 1.35, 0.65, 0.85, 3.7])

        with limit_label_col:
            st.markdown("**件数**")

        with limit_col:
            result_limit = st.selectbox(
                "表示件数",
                [5, 10, 20],
                key="journal_result_limit",
                label_visibility="collapsed"
            )

        if (
            st.session_state.get("last_journal_search_limit")
            != result_limit
        ):
            search_params = st.session_state.get(
                "last_journal_search_params",
                {
                    "keyword": st.session_state.get("keyword_input", ""),
                    "dept": dept if dept else None,
                    "amount": amount,
                }
            )
            st.session_state.results = search(
                records,
                search_params.get("keyword", ""),
                search_params.get("dept"),
                search_params.get("amount"),
                freq,
                limit=result_limit
            )
            st.session_state["last_journal_search_limit"] = result_limit
            st.session_state.pop("selected_candidate_index", None)
            st.session_state["selected_candidate_no"] = 1
            st.session_state["candidate_number_select_input"] = 1

        results = st.session_state.results
        result_count_message.success(f"{len(results)}件ヒット")

        if (
            st.session_state.get("selected_candidate_no", 1) < 1
            or st.session_state.get("selected_candidate_no", 1) > len(results)
        ):
            st.session_state.pop("selected_candidate_no", None)

        current_selected_index = st.session_state.get(
            "selected_candidate_index"
        )
        default_candidate_no = (
            current_selected_index + 1
            if (
                current_selected_index is not None
                and 0 <= current_selected_index < len(results)
            )
            else st.session_state.get("selected_candidate_no", 1)
        )
        if "candidate_number_select_input" not in st.session_state:
            st.session_state[
                "candidate_number_select_input"
            ] = default_candidate_no
        elif (
            st.session_state.get("candidate_number_select_input", 1) < 1
            or st.session_state.get(
                "candidate_number_select_input",
                1
            ) > len(results)
        ):
            st.session_state[
                "candidate_number_select_input"
            ] = default_candidate_no

        def apply_candidate_selection():
            selected_no = int(
                st.session_state.get(
                    "candidate_number_select_input",
                    default_candidate_no,
                )
            )
            selected_index = selected_no - 1
            if 0 <= selected_index < len(results):
                st.session_state[
                    "selected_candidate_index"
                ] = selected_index
                st.session_state[
                    "selected_candidate_no"
                ] = selected_no
                st.session_state["candidate_select_message"] = (
                    f"候補{selected_no}を編集対象にしました。"
                )
            else:
                st.session_state[
                    "candidate_select_message"
                ] = "候補番号が範囲外です。"

        with label_col:
            st.markdown("**候補番号**")

        with input_col:
            st.number_input(
                "候補番号",
                min_value=1,
                max_value=len(results),
                value=default_candidate_no,
                step=1,
                key="candidate_number_select_input",
                label_visibility="collapsed",
                on_change=apply_candidate_selection,
            )

        with button_col:
            st.button(
                "編集対象にする",
                key="apply_candidate_number_select",
                on_click=apply_candidate_selection,
            )

        candidate_select_message = st.session_state.pop(
            "candidate_select_message",
            "",
        )
        if candidate_select_message:
            if "範囲外" in candidate_select_message:
                st.warning(candidate_select_message)
            else:
                st.success(candidate_select_message)

        selected_candidate_index = st.session_state.get(
            "selected_candidate_index"
        )
        if (
            selected_candidate_index is not None
            and 0 <= selected_candidate_index < len(results)
        ):
            selected_score, selected_rec, _ = results[
                selected_candidate_index
            ]
            if isinstance(selected_rec, dict) and "rows" in selected_rec:
                selected_matched_row = get_matched_amount_row(selected_rec)
                selected_rows = split_journal(selected_rec["rows"])
                if selected_rows:
                    selected_first_row = selected_rows[0]

                    st.subheader("現在編集中の仕訳")

                    st.dataframe(
                        pd.DataFrame([{
                            "候補": f"{selected_candidate_index + 1}【選択中】",
                            "日付": f"{process_date_obj:%Y/%m/%d}",
                            "借方": format_account_name_with_code(
                                selected_first_row.get(COL_DEBIT, "")
                            ),
                            "貸方": format_account_name_with_code(
                                selected_first_row.get(COL_CREDIT, "")
                            ),
                            "金額": f"{get_voucher_total(selected_rows):,}",
                            "摘要": selected_first_row.get(COL_SUMMARY, ""),
                        }]),
                        hide_index=True,
                        use_container_width=True
                    )
                    st.caption("下の候補詳細で編集・登録してください。")
                    if selected_matched_row:
                        st.caption(
                            "検索一致行は下の候補詳細に別表示しています。"
                        )

        selected_candidate_index = st.session_state.get(
            "selected_candidate_index"
        )
        display_indices = list(range(len(results)))
        if (
            selected_candidate_index is not None
            and 0 <= selected_candidate_index < len(results)
        ):
            display_indices = [selected_candidate_index] + [
                display_index
                for display_index in display_indices
                if display_index != selected_candidate_index
            ]

        for display_index in display_indices:

            score, rec, score_detail = results[display_index]
            idx = display_index + 1
    
            if not isinstance(rec, dict):
                continue
    
            if "rows" not in rec:
                continue
    
            rows = split_journal(rec["rows"])
    
            doc_id = f"{idx}_{id(rec)}"
            is_selected_candidate = selected_candidate_index == display_index
            selected_marker = "【選択中】" if is_selected_candidate else ""
            matched_amount_row = get_matched_amount_row(rec)
            show_voucher_block = should_show_voucher_block(
                rec,
                matched_amount_row
            )
            recommendation_label = (
                "候補"
                if show_voucher_block
                else "推奨"
            )
            matched_row_title = format_matched_row_title(
                matched_amount_row
            )
            amount_match_detail = extract_amount_match_detail(score_detail)
    
            if matched_row_title:
                summary = (
                    f"候補 {idx}{selected_marker} / スコア {score} "
                    f"{matched_row_title}"
                )
            else:
                summary = (
                    f"候補 {idx}{selected_marker} / スコア {score} "
                    f"{rows[0].get('摘要','')}"
                    f"　{len(rows)}行"
                    f"　¥{get_voucher_total(rows):,}"
                )
                if amount_match_detail:
                    summary += f"　金額一致/近似行: {amount_match_detail}"
    
            with st.expander(
                summary,
                expanded=is_selected_candidate
            ):

                if is_selected_candidate:
                    st.info(
                        "現在この候補が編集対象として選択されています。"
                    )

                if matched_amount_row:
                    st.markdown("**検索一致行**")
                    st.dataframe(
                        pd.DataFrame([
                            build_matched_row_display(matched_amount_row)
                        ]),
                        hide_index=True,
                        use_container_width=True
                    )
                    if (
                        is_excluded_account(
                            matched_amount_row.get("debit", "")
                        )
                        or is_excluded_account(
                            matched_amount_row.get("credit", "")
                        )
                    ):
                        st.caption(
                            "この検索一致行には「資金複合 / 諸口」が"
                            "含まれています。編集欄では直接選択できない"
                            "ため、実際に登録する相手科目へ変更して"
                            "ください。"
                        )

                if show_voucher_block:
                    st.info(
                        "この候補は「資金複合 / 諸口」を含む"
                        "複合仕訳です。編集欄では資金複合 / 諸口を"
                        "直接選択できないため、下の同一伝票ブロックを"
                        "確認し、実際に登録する相手科目へ変更して"
                        "ください。"
                    )
                    st.markdown("**同一伝票ブロック（参考）**")
                    st.dataframe(
                        pd.DataFrame(
                            build_voucher_block_display(
                                rec.get("rows", [])
                            )
                        ),
                        hide_index=True,
                        use_container_width=True
                    )
                    st.caption(
                        "表示のみです。登録内容やCSV出力値は、"
                        "下の編集欄で確定した内容が使われます。"
                    )
    
                with st.expander("検索理由"):
    
                    full_match_count = len([
                        d for d in score_detail
                        if "完全一致" in d
                    ])
    
                    partial_match_count = len([
                        d for d in score_detail
                        if "部分一致" in d
                    ])
    
                    if full_match_count:
                        st.write(
                            f"✅ 完全一致 {full_match_count}件"
                        )
    
                    if partial_match_count:
                        st.write(
                            f"✅ 部分一致 {partial_match_count}件"
                        )
    
                    if any(
                        "部門一致" in d
                        for d in score_detail
                    ):
                        st.write("✅ 部門一致")
    
                    if any(
                        "複数キーワード一致" in d
                        for d in score_detail
                    ):
                        st.write("✅ 複数キーワード一致")

                    if matched_amount_row:
                        st.write(
                            "✅ 金額一致/近似行: "
                            f"{format_matched_row_title(matched_amount_row)}"
                        )
                    elif amount_match_detail:
                        st.write(
                            f"✅ 金額一致/近似行: {amount_match_detail}"
                        )
    
                st.divider()
    
                edited_rows = []

                d_sum = 0
                c_sum = 0
                entered_amounts_valid = True

                for r_idx, r in enumerate(rows):

                    debit_account = get_account_name(
                        r.get("借方科目", "")
                    )

                    credit_account = get_account_name(
                        r.get("貸方科目", "")
                    )
                    amount_value = to_int(
                        r.get(COL_DEBIT_AMOUNT, 0)
                    )

                    row_summary = (
                        f"{r_idx+1}行目 "
                        f"🔵[借] {debit_account} "
                        f"/ "
                        f"🔴[貸] {credit_account} "
                        f"/ ¥{amount_value:,}"
                    )

                    with st.expander(
                        row_summary,
                        expanded=is_selected_candidate
                    ):

                        col1, col2 = st.columns(2)

                        # =====================================
                        # 借方 / 貸方
                        # =====================================
                        summary_for_suggestions = " ".join([
                            str(r.get(COL_SUMMARY, "") or ""),
                            str(r.get("摘要", "") or ""),
                            str(r.get("伝票摘要", "") or "")
                        ])
                        debit_key = f"d_{doc_id}_{r_idx}"
                        credit_key = f"c_{doc_id}_{r_idx}"

                        with col1:

                            default_debit = r.get(
                                COL_DEBIT,
                                ""
                            )
                            block_debit_candidates = (
                                infer_block_replacement_accounts(
                                    rec.get("rows", rows),
                                    r,
                                    "debit"
                                )
                            )
                            current_credit = st.session_state.get(
                                credit_key,
                                r.get(COL_CREDIT, "")
                            )
                            debit_options, recommended_debits = (
                                build_account_select_options(
                                    records,
                                    summary_for_suggestions,
                                    current_credit,
                                    sub_account=r.get(COL_DEBIT_SUB, ""),
                                    current_account=default_debit,
                                    priority_accounts=block_debit_candidates
                                )
                            )
                            debit_label_accounts = get_account_label_set(
                                recommended_debits,
                                (
                                    block_debit_candidates
                                    if show_voucher_block
                                    else None
                                )
                            )

                            debit = st.selectbox(
                                "借方",
                                debit_options,
                                index=(
                                    debit_options.index(default_debit)
                                    if (
                                        default_debit in debit_options
                                        and not is_excluded_account(
                                            default_debit
                                        )
                                    )
                                    else 0
                                ),
                                key=debit_key,
                                format_func=(
                                    lambda account,
                                    recommended=debit_label_accounts,
                                    label=recommendation_label:
                                    format_recommended_account(
                                        account,
                                        recommended,
                                        label=label
                                    )
                                )
                            )
                            if is_excluded_account(default_debit):
                                st.caption(
                                    "元データの科目が資金複合/諸口のため、"
                                    "選択可能な科目へ変更してください"
                                )

                        with col2:

                            default_credit = r.get(
                                COL_CREDIT,
                                ""
                            )
                            block_credit_candidates = (
                                infer_block_replacement_accounts(
                                    rec.get("rows", rows),
                                    r,
                                    "credit"
                                )
                            )
                            credit_options, recommended_credits = (
                                build_account_select_options(
                                    records,
                                    summary_for_suggestions,
                                    debit,
                                    sub_account=r.get(COL_CREDIT_SUB, ""),
                                    current_account=default_credit,
                                    priority_accounts=block_credit_candidates
                                )
                            )
                            credit_label_accounts = get_account_label_set(
                                recommended_credits,
                                (
                                    block_credit_candidates
                                    if show_voucher_block
                                    else None
                                )
                            )

                            credit = st.selectbox(
                                "貸方",
                                credit_options,
                                index=(
                                    credit_options.index(default_credit)
                                    if (
                                        default_credit in credit_options
                                        and not is_excluded_account(
                                            default_credit
                                        )
                                    )
                                    else 0
                                ),
                                key=credit_key,
                                format_func=(
                                    lambda account,
                                    recommended=credit_label_accounts,
                                    label=recommendation_label:
                                    format_recommended_account(
                                        account,
                                        recommended,
                                        label=label
                                    )
                                )
                            )
                            if is_excluded_account(default_credit):
                                st.caption(
                                    "元データの科目が資金複合/諸口のため、"
                                    "選択可能な科目へ変更してください"
                                )

                    # =====================================
                    # 補助
                    # =====================================
                    col3, col4 = st.columns(2)

                    with col3:

                        default_ds = r.get(
                            COL_DEBIT_SUB,
                            ""
                        )
                        debit_sub_options = build_sub_options(default_ds)

                        debit_sub = st.selectbox(
                            "借方補助",
                            debit_sub_options,
                            index=(
                                debit_sub_options.index(default_ds)
                                if default_ds in debit_sub_options
                                else 0
                            ),
                            key=f"ds_{doc_id}_{r_idx}"
                        )

                    with col4:

                        default_cs = r.get(
                            COL_CREDIT_SUB,
                            ""
                        )
                        credit_sub_options = build_sub_options(default_cs)

                        credit_sub = st.selectbox(
                            "貸方補助",
                            credit_sub_options,
                            index=(
                                credit_sub_options.index(default_cs)
                                if default_cs in credit_sub_options
                                else 0
                            ),
                            key=f"cs_{doc_id}_{r_idx}"
                        )

                    # =====================================
                    # 金額
                    # =====================================
                    suggest = get_amount_suggestions(
                        records,
                        debit,
                        credit
                    )

                    st.caption(f"過去金額: ¥{amount_value:,}")

                    if suggest:
                        st.caption(
                            f"平均金額: ¥{suggest['avg']:,}"
                        )

                    default_amt = (
                        amount
                        if len(rows) == 1
                        and amount is not None
                        and amount > 0
                        else None
                    )

                    amount_col, summary_col = st.columns([1, 2])

                    with amount_col:
                        amt = st.number_input(
                            "金額",
                            min_value=0,
                            value=default_amt,
                            step=1,
                            placeholder="今回の金額",
                            key=f"amt_{doc_id}_{r_idx}"
                        )

                    # =====================================
                    # 摘要
                    # =====================================
                    with summary_col:
                        memo = st.text_input(
                            "摘要",
                            value=r.get(COL_SUMMARY, ""),
                            key=f"m_{doc_id}_{r_idx}"
                        )

                    if amt is None or amt <= 0:
                        entered_amounts_valid = False
                        registered_amount = 0
                    else:
                        registered_amount = int(amt)

                    d_sum += registered_amount
                    c_sum += registered_amount

                    new_row = copy.deepcopy(r)

                    new_row[COL_DATE] = process_date

                    new_row[COL_DEBIT] = debit
                    new_row[COL_CREDIT] = credit

                    new_row[COL_DEBIT_SUB] = debit_sub
                    new_row[COL_CREDIT_SUB] = credit_sub

                    new_row[COL_DEBIT_AMOUNT] = (
                        str(registered_amount)
                        if registered_amount > 0
                        else ""
                    )
                    new_row[COL_CREDIT_AMOUNT] = (
                        str(registered_amount)
                        if registered_amount > 0
                        else ""
                    )

                    new_row[COL_SUMMARY] = memo

                    edited_rows.append(new_row)

                st.divider()

                # =====================================
                # 未来日付
                # =====================================
                if process_date_obj > datetime.today().date():
                    st.warning("⚠️ 未来日付")

                # =====================================
                # 登録
                # =====================================
                if st.button(
                    "この内容で登録",
                    key=f"save_{doc_id}",
                    type="primary"
                ):

                    if not entered_amounts_valid:

                        st.warning("今回の金額を入力してください")

                    elif d_sum != c_sum:

                        st.error(
                            "借貸が一致していません。"
                            f"借方¥{d_sum:,} / 貸方¥{c_sum:,}"
                        )

                    else:

                        st.session_state.confirmed.append(
                            copy.deepcopy(edited_rows)
                        )

                        st.session_state[
                            f"next_search_ready_{doc_id}"
                        ] = True

                if st.session_state.get(f"next_search_ready_{doc_id}"):
                    st.success(
                        "登録しました。次の仕訳検索へ進めます。"
                    )
                    st.button(
                        "次の仕訳を検索する",
                        key=f"reset_for_next_journal_{doc_id}",
                        type="secondary",
                        on_click=reset_for_next_journal_search
                    )
    
    # =========================================
    # 登録済
    # =========================================
    if st.session_state.confirmed:
    
        st.divider()
    
        st.header("📦 登録済仕訳（編集可能）")
    
        for doc_idx, doc in enumerate(
            st.session_state.confirmed
        ):
    
            with st.expander(f"伝票 {doc_idx+1}"):
    
                edited_doc = []
    
                for row_idx, r in enumerate(doc):
    
                    st.markdown(f"### 行 {row_idx+1}")
    
                    col1, col2 = st.columns(2)
    
                    with col1:
    
                        debit = st.selectbox(
                            "借方",
                            account_master,
                            index=(
                                account_master.index(
                                    r[COL_DEBIT]
                                )
                                if r[COL_DEBIT] in account_master
                                else 0
                            ),
                            key=f"conf_d_{doc_idx}_{row_idx}"
                        )
    
                    with col2:
    
                        credit = st.selectbox(
                            "貸方",
                            account_master,
                            index=(
                                account_master.index(
                                    r[COL_CREDIT]
                                )
                                if r[COL_CREDIT] in account_master
                                else 0
                            ),
                            key=f"conf_c_{doc_idx}_{row_idx}"
                        )
    
                    col3, col4 = st.columns(2)
    
                    with col3:

                        default_ds = r.get(
                            COL_DEBIT_SUB,
                            ""
                        )
                        debit_sub_options = build_sub_options(default_ds)
    
                        debit_sub = st.selectbox(
                            "借方補助",
                            debit_sub_options,
                            index=(
                                debit_sub_options.index(default_ds)
                                if default_ds in debit_sub_options
                                else 0
                            ),
                            key=f"conf_ds_{doc_idx}_{row_idx}"
                        )
    
                    with col4:

                        default_cs = r.get(
                            COL_CREDIT_SUB,
                            ""
                        )
                        credit_sub_options = build_sub_options(default_cs)
    
                        credit_sub = st.selectbox(
                            "貸方補助",
                            credit_sub_options,
                            index=(
                                credit_sub_options.index(default_cs)
                                if default_cs in credit_sub_options
                                else 0
                            ),
                            key=f"conf_cs_{doc_idx}_{row_idx}"
                        )
    
                    amount_col, summary_col = st.columns([1, 2])

                    with amount_col:

                        amt = st.number_input(
                            "金額",
                            value=to_int(
                                r["借方金額"]
                            ),
                            key=f"conf_amt_{doc_idx}_{row_idx}"
                        )
    
                    with summary_col:

                        memo = st.text_input(
                            "摘要",
                            value=r.get(COL_SUMMARY, ""),
                            key=f"conf_m_{doc_idx}_{row_idx}"
                        )
    
                    new_row = copy.deepcopy(r)
    
                    new_row[COL_DEBIT] = debit
                    new_row[COL_CREDIT] = credit
    
                    new_row[COL_DEBIT_SUB] = debit_sub
                    new_row[COL_CREDIT_SUB] = credit_sub
    
                    new_row[COL_DEBIT_AMOUNT] = str(amt)
                    new_row[COL_CREDIT_AMOUNT] = str(amt)
    
                    new_row[COL_SUMMARY] = memo
    
                    edited_doc.append(new_row)
    
                colA, colB = st.columns(2)
    
                # =====================================
                # 更新保存
                # =====================================
                with colA:
    
                    if st.button(
                        "更新保存",
                        key=f"update_{doc_idx}"
                    ):
    
                        st.session_state.confirmed[
                            doc_idx
                        ] = edited_doc
    
                        st.success("更新しました")
    
                        st.rerun()
    
                # =====================================
                # 削除
                # =====================================
                with colB:
    
                    if st.button(
                        "削除",
                        key=f"delete_{doc_idx}"
                    ):
    
                        st.session_state.confirmed.pop(
                            doc_idx
                        )
    
                        st.rerun()
    
        # =========================================
        # 出力
        # =========================================
        st.divider()
    
        st.header("📄 出力")
    
        all_rows = []
    
        for doc in st.session_state.confirmed:
            all_rows.extend(doc)
    
        # =====================================
        # 入力用Excel
        # =====================================
        st.subheader("入力用Excel（簡易仕訳帳・印刷用）")
        st.caption(
            "エプソンへ手入力・修正するための印刷用一覧です。45列インポート形式ではありません。"
        )

        input_rows = build_input_csv_rows(
            st.session_state.get("confirmed", [])
        )
        input_df = pd.DataFrame(
            input_rows,
            columns=INPUT_CSV_COLUMNS
        ).fillna("")

        st.dataframe(
            input_df,
            use_container_width=True
        )

        input_excel = build_input_journal_excel(input_df)
        input_excel_filename = (
            "input_journal_print_"
            f"{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
    
        input_save_col, input_download_col = st.columns([1, 1])

        with input_save_col:
            if st.button(
                "共有フォルダへ保存",
                key="save_input_excel_to_export_dir",
                type="primary"
            ):
                saved, message = save_file_to_export_dir(
                    input_excel,
                    input_excel_filename,
                    st.session_state.get("csv_export_dir", "")
                )
                if saved:
                    st.success(message)
                else:
                    st.warning(message)

        with input_download_col:
            st.download_button(
                "ブラウザでダウンロード",
                data=input_excel,
                file_name=input_excel_filename,
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                )
            )
    
        # =====================================
        # エプソンCSV
        # =====================================
        epson_company_name = str(
            st.session_state.get("company_name", "")
        ).strip()

        if not epson_company_name:
            st.warning(
                "入力会社が空欄です。エプソン取込CSVの入力会社列も空欄になります。"
            )

        epson_rows = build_epson_rows(
            all_rows,
            epson_company_name
        )
    
        epson_df = pd.DataFrame(
            epson_rows,
            columns=EPSON_COLUMNS
        ).fillna("")
    
        epson_csv = epson_df.to_csv(
            index=False
        ).encode("cp932")
        epson_filename = (
            "epson_output_"
            f"{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        )

        if "epson_export_success" in st.session_state:
            st.success(
                st.session_state.pop("epson_export_success")
            )
    
        st.subheader("エプソン取込CSV")
        st.caption("登録済み仕訳をエプソン取込形式で保存します。")
        epson_save_col, epson_download_col = st.columns([1, 1])

        with epson_save_col:
            if st.button(
                "共有フォルダへ保存",
                key="save_epson_csv_to_export_dir",
                type="primary"
            ):
                saved, message = save_csv_to_export_dir(
                    epson_csv,
                    epson_filename,
                    st.session_state.get("csv_export_dir", "")
                )
                if saved:
                    st.success(message)
                else:
                    st.warning(message)

        with epson_download_col:
            st.download_button(
                "ブラウザでダウンロード",
                epson_csv,
                epson_filename,
                on_click=save_exported_journals,
                args=(epson_rows,)
            )
    

elif mode == "未収消込":

    st.header("未収消込")

    if "receivable_import_success" in st.session_state:
        st.success(
            st.session_state.pop("receivable_import_success")
        )

    if "receivable_import_key" not in st.session_state:
        st.session_state.receivable_import_key = 0

    with st.expander("未収一覧CSV取込"):

        format_col, upload_col = st.columns([1, 2])

        with format_col:
            import_format = st.selectbox(
                "取込形式",
                [
                    "標準未収CSV形式",
                    "請求一覧Excel形式"
                ],
                key=(
                    "receivable_import_format_"
                    f"{st.session_state.receivable_import_key}"
                )
            )

        import_preview = None
        import_errors = pd.DataFrame()
        excluded_duplicate_count = 0

        if import_format == "標準未収CSV形式":

            with upload_col:
                uploaded_receivables = st.file_uploader(
                    "未収一覧CSV",
                    type=["csv"],
                    key=(
                        "receivable_import_csv_"
                        f"{st.session_state.receivable_import_key}"
                    )
                )

            st.markdown(
                "標準未収CSV 必須列：\n\n"
                "取引先,請求日,請求額,残高,摘要"
            )
            st.markdown(
                "任意列：\n\n"
                "入金予定日,未収科目,未収補助,部門"
            )
            st.markdown("サンプル：")
            st.code(
                "取引先,請求日,入金予定日,未収科目,未収補助,部門,摘要,請求額,残高\n"
                "サンプル運送株式会社,2026-06-30,2026-07-31,未収運賃,"
                "サンプル運送株式会社,総務課,6月分運賃,50000,50000",
                language="csv"
            )
            st.markdown(
                "注意：\n"
                "- サンプルの企業名は架空です\n"
                "- 日付は YYYY-MM-DD / YYYY/MM/DD / YYYYMMDD に対応\n"
                "- 金額はカンマなし推奨\n"
                "- 任意列がない場合は既定値で補完します\n"
                "- 請求額と残高は必須です"
            )

        else:

            with upload_col:
                uploaded_receivables = st.file_uploader(
                    "請求一覧Excel",
                    type=["xlsx", "xls"],
                    key=(
                        "receivable_import_excel_"
                        f"{st.session_state.receivable_import_key}"
                    )
                )

            date_col, account_col, dept_col = st.columns([1, 1.3, 1.3])

            with date_col:
                invoice_date = st.date_input(
                    "請求日",
                    key=(
                        "company_invoice_date_"
                        f"{st.session_state.receivable_import_key}"
                    )
                )

            with account_col:
                default_receivable_account = st.selectbox(
                    "既定の未収科目",
                    account_master,
                    index=(
                        account_master.index("未収運賃")
                        if "未収運賃" in account_master
                        else 0
                    ),
                    key=(
                        "company_receivable_account_"
                        f"{st.session_state.receivable_import_key}"
                    )
                )

            with dept_col:
                import_department = st.selectbox(
                    "部門",
                    [""] + department_master,
                    key=(
                        "company_receivable_department_"
                        f"{st.session_state.receivable_import_key}"
                    )
                )

            specify_payment_due_date = st.checkbox(
                "入金予定日を指定する",
                key=(
                    "specify_payment_due_date_"
                    f"{st.session_state.receivable_import_key}"
                )
            )

            if specify_payment_due_date:
                payment_due_date = st.date_input(
                    "入金予定日",
                    key=(
                        "company_payment_due_date_"
                        f"{st.session_state.receivable_import_key}"
                    )
                )
            else:
                payment_due_date = None
                st.caption(
                    "入金予定日は請求日の翌月末で設定します"
                )

        if uploaded_receivables is not None:

            try:
                if import_format == "標準未収CSV形式":
                    try:
                        uploaded_receivables.seek(0)
                        source_receivables = pd.read_csv(
                            uploaded_receivables,
                            dtype=str,
                            encoding="utf-8-sig"
                        ).fillna("")
                    except UnicodeDecodeError:
                        uploaded_receivables.seek(0)
                        source_receivables = pd.read_csv(
                            uploaded_receivables,
                            dtype=str,
                            encoding="cp932"
                        ).fillna("")

                    import_preview, import_errors = (
                        normalize_standard_receivable_csv(
                            source_receivables
                        )
                    )

                else:
                    excel_data = uploaded_receivables.getvalue()
                    excel_file = pd.ExcelFile(
                        io.BytesIO(excel_data)
                    )
                    sheet_name = (
                        "プリント用"
                        if "プリント用" in excel_file.sheet_names
                        else excel_file.sheet_names[0]
                    )
                    raw_billing_df = pd.read_excel(
                        io.BytesIO(excel_data),
                        sheet_name=sheet_name,
                        header=None,
                        dtype=object
                    )
                    st.caption(f"読込シート: {sheet_name}")

                    standard_source, conversion_errors = (
                        convert_company_billing_excel(
                            raw_billing_df,
                            invoice_date,
                            payment_due_date,
                            default_receivable_account,
                            import_department
                        )
                    )
                    import_preview, validation_errors = (
                        normalize_standard_receivable_csv(
                            standard_source
                        )
                    )
                    company_duplicate_columns = [
                        "コード",
                        "得意先名",
                        "請求日",
                        "請求金額",
                        "未収科目",
                        "未収補助"
                    ]
                    import_preview, duplicate_errors = (
                        exclude_duplicate_receivables(
                            import_preview,
                            company_duplicate_columns
                        )
                    )
                    excluded_duplicate_count = len(
                        duplicate_errors
                    )
                    import_errors = pd.concat(
                        [
                            conversion_errors,
                            validation_errors,
                            duplicate_errors.drop(
                                columns=["未収ID"],
                                errors="ignore"
                            )
                        ],
                        ignore_index=True
                    )

                if excluded_duplicate_count:
                    st.warning(
                        "取り込み済みの請求を"
                        f"{excluded_duplicate_count}件除外しました"
                    )

                if not import_errors.empty:
                    st.error(
                        f"取込対象外の行が{len(import_errors)}件あります"
                    )
                    st.dataframe(
                        import_errors,
                        use_container_width=True
                    )

                st.write("取込プレビュー")
                st.dataframe(
                    import_preview.drop(
                        columns=["未収ID"],
                        errors="ignore"
                    ),
                    use_container_width=True
                )

                if import_preview.empty:
                    st.warning("取り込める明細がありません")

                else:
                    st.caption("プレビューの明細を未収一覧へ追加します。")

                if not import_preview.empty and st.button(
                    "未収一覧へ取り込む",
                    key=(
                        "append_receivables_"
                        f"{st.session_state.receivable_import_key}"
                    ),
                    type="primary"
                ):
                    duplicate_columns = (
                        [
                            "コード",
                            "得意先名",
                            "請求日",
                            "請求金額",
                            "未収科目",
                            "未収補助"
                        ]
                        if import_format == "請求一覧Excel形式"
                        else None
                    )
                    imported_count, duplicate_count = (
                        append_standard_receivables(
                            import_preview,
                            duplicate_columns=duplicate_columns
                        )
                    )

                    if imported_count:
                        message = (
                            f"未収一覧へ{imported_count}件取り込みました"
                        )
                    else:
                        message = "追加対象の未収明細はありません"

                    if excluded_duplicate_count or duplicate_count:
                        message += (
                            "（重複"
                            f"{excluded_duplicate_count + duplicate_count}"
                            "件を除外）"
                        )

                    st.session_state[
                        "receivable_import_success"
                    ] = message
                    st.session_state.receivable_import_key += 1
                    st.rerun()

            except ImportError as e:
                if (
                    import_format == "請求一覧Excel形式"
                    and "openpyxl" in str(e).casefold()
                ):
                    st.error(
                        "Excelファイルの読み込みに必要な "
                        "openpyxl がインストールされていません。"
                        "\n\n以下を実行してください："
                        "\n\n`pip install openpyxl`"
                    )
                else:
                    st.error(
                        f"未収一覧ファイルを読み込めません: {e}"
                    )
            except ValueError as e:
                st.error(str(e))
            except Exception as e:
                st.error(f"未収一覧ファイルを読み込めません: {e}")

    if "receivable_cleanup_success" in st.session_state:
        st.success(
            st.session_state.pop("receivable_cleanup_success")
        )

    cleanup_df = load_receivables()
    cleanup_balances = pd.to_numeric(
        cleanup_df["残高"].astype(str).str.replace(",", ""),
        errors="coerce"
    )
    cleanup_mask = (
        cleanup_df["ステータス"].astype(str).str.strip().eq("完了")
        | cleanup_balances.le(0)
    )
    cleanup_count = int(cleanup_mask.sum())

    with st.expander("未収台帳の整理"):
        st.write(
            "完了済みまたは残高0の未収が"
            f"{cleanup_count}件あります"
        )
        st.caption(
            "整理すると、current.csv から完了済み未収を除外します。"
            "消込履歴と検索DBは削除されません。"
        )

        if st.button(
            "完了済み未収を整理する",
            key="organize_completed_receivables",
            disabled=cleanup_count == 0
        ):
            organized_count = organize_completed_receivables()
            st.session_state["receivable_cleanup_success"] = (
                f"完了済み未収を{organized_count}件整理しました"
            )
            st.rerun()

    payment_accounts = load_payment_accounts()

    if "account_master_success" in st.session_state:
        for message in st.session_state.pop(
            "account_master_success"
        ):
            st.success(message)

    default_receipt_account = (
        "普通預金"
        if "普通預金" in payment_accounts
        else payment_accounts[0]
    )

    if (
        "receipt_account" not in st.session_state
        or st.session_state.receipt_account
        not in payment_accounts
    ):
        st.session_state.receipt_account = default_receipt_account

    if "receivable_success" in st.session_state:
        st.success(
            st.session_state.pop("receivable_success")
        )

    receivables_df = load_receivables()

    if receivables_df.empty:

        st.info("未収データがありません")

    else:

        receivables_df = receivables_df.copy()

        receivables_df = receivables_df[
            receivables_df["得意先名"].astype(str).str.strip() != ""
        ].copy()

        receivables_df["残高"] = pd.to_numeric(
            receivables_df["残高"].str.replace(",", ""),
            errors="coerce"
        ).fillna(0).astype(int)

        receivables_df = receivables_df[
            (receivables_df["残高"] > 0)
            &
            (receivables_df["ステータス"] != "完了")
        ].copy()

        if receivables_df.empty:

            st.info("未収データがありません")

        else:

            balance_df = (
                receivables_df
                .groupby(
                    "得意先名",
                    as_index=False
                )
                .agg(
                    残高=("残高", "sum"),
                    件数=("残高", "size")
                )
                .rename(
                    columns={"得意先名": "取引先"}
                )
                .sort_values(
                    "残高",
                    ascending=False
                )
            )

            st.divider()
            st.subheader("① 取込サマリー（参照専用）")
            st.info(
                "取引先別の未収残高と件数を確認できます。消込処理は下の『② 消込作業エリア』から行ってください。"
            )

            balance_summary_df = balance_df[
                ["取引先", "残高", "件数"]
            ].copy()
            balance_summary_df["残高"] = balance_summary_df[
                "残高"
            ].map(lambda value: f"{int(value):,}")

            st.table(balance_summary_df)
            st.divider()

            st.subheader("② 消込作業エリア")
            st.caption(
                "処理したい取引先を開き、確認情報、入金入力、候補表示の順に進めてください。"
            )

            for customer_idx, (_, customer) in enumerate(
                balance_df.iterrows()
            ):

                customer_name = customer["取引先"]
                customer_count = int(customer["件数"])
                customer_balance = int(customer["残高"])
                customer_balance_text = f"{customer_balance:,}"

                with st.expander(
                    (
                        "消込作業を開始："
                        f"{customer_name}"
                        f"（{customer_count}件 / {customer_balance_text}円）"
                    ),
                    expanded=(
                        st.session_state.get(
                            "active_receivable_customer"
                        ) == customer_name
                        or st.session_state.get(
                            "active_receivable_customer_idx"
                        ) == customer_idx
                        or st.session_state.get(
                            "open_receivable_customer"
                        ) == customer_name
                    )
                ):

                    detail_df = receivables_df[
                        receivables_df["得意先名"] == customer_name
                    ].copy()

                    if "請求日" not in detail_df.columns:
                        detail_df["請求日"] = ""

                    detail_df["請求金額"] = pd.to_numeric(
                        detail_df["請求金額"].str.replace(",", ""),
                        errors="coerce"
                    ).fillna(0).astype(int)

                    detail_df["入金済額"] = (
                        detail_df["請求金額"]
                        - detail_df["残高"]
                    )

                    detail_display_columns = [
                        "コード",
                        "請求日",
                        "請求金額",
                        "入金済額",
                        "残高",
                        "ステータス",
                        "未収科目",
                        "未収補助",
                        "部門"
                    ]

                    detail_storage_columns = (
                        ["未収ID"]
                        + detail_display_columns
                        + [
                            column
                            for column in ["摘要"]
                            if column in detail_df.columns
                        ]
                    )
                    detail_df = detail_df[
                        detail_storage_columns
                    ]

                    st.info(
                        "\n".join([
                            "現在処理中：",
                            f"取引先：{customer_name}",
                            f"対象件数：{customer_count}件",
                            f"残高合計：{customer_balance_text}円",
                        ])
                    )

                    st.caption("確認情報")
                    st.dataframe(
                        detail_df[detail_display_columns],
                        use_container_width=True
                    )

                    payment_amount_key = (
                        "payment_amount_text_"
                        f"{customer_idx}_{customer_name}"
                    )
                    payment_submission_key = (
                        f"{customer_idx}_{customer_name}"
                    )
                    payment_candidates_key = (
                        f"payment_candidates_{customer_idx}"
                    )

                    with st.form(
                        key=f"payment_form_{customer_idx}_{customer_name}"
                    ):

                        st.caption("入金入力")
                        (
                            payment_date_col,
                            payment_amount_col,
                            receipt_account_col
                        ) = st.columns([1, 1, 1.4])

                        with payment_date_col:
                            payment_date = st.date_input(
                                "入金日",
                                key=f"payment_date_{customer_idx}"
                            )

                        with payment_amount_col:
                            payment_amount_text = st.text_input(
                                "入金額",
                                placeholder="0",
                                key=payment_amount_key
                            )

                        with receipt_account_col:
                            receipt_account = st.selectbox(
                                "入金科目",
                                payment_accounts,
                                index=(
                                    payment_accounts.index(
                                        st.session_state.receipt_account
                                    )
                                ),
                                key=f"receipt_account_{customer_idx}"
                            )

                        st.caption(
                            "入金額を入力し、Enterキーでも候補を表示できます。"
                        )
                        payment_preview_submitted = (
                            st.form_submit_button(
                                "入金候補を表示",
                                type="primary"
                            )
                        )

                    with st.expander("科目候補の追加"):

                        st.info(
                            "入金科目がない場合は、通常仕訳画面の左サイドバーにある「科目マスター管理」から追加してください。"
                        )

                    if payment_preview_submitted:
                        keep_receivable_customer_open(customer_name)
                        st.session_state[
                            "active_receivable_customer_idx"
                        ] = customer_idx
                        st.session_state[
                            "submitted_payment_amount"
                        ] = st.session_state.get(
                            payment_amount_key,
                            payment_amount_text
                        )
                        st.session_state[
                            "submitted_payment_account"
                        ] = receipt_account
                        st.session_state[
                            "submitted_payment_date"
                        ] = payment_date
                        st.session_state[
                            "pending_receivable_submission_key"
                        ] = payment_submission_key
                        st.rerun()

                    process_payment_submission = (
                        st.session_state.get(
                            "pending_receivable_submission_key"
                        ) == payment_submission_key
                    )
                    submitted_payment_amount = st.session_state.get(
                        "submitted_payment_amount",
                        st.session_state.get(
                            payment_amount_key,
                            payment_amount_text
                        )
                    )
                    submitted_payment_account = st.session_state.get(
                        "submitted_payment_account",
                        receipt_account
                    )
                    submitted_payment_date = st.session_state.get(
                        "submitted_payment_date",
                        payment_date
                    )
                    parsed_payment_amount = (
                        parse_receivable_payment_amount(
                            submitted_payment_amount
                        )
                    )

                    if (
                        process_payment_submission
                        and parsed_payment_amount is None
                    ):
                        st.warning("入金額を入力してください")
                        st.session_state.pop(
                            "pending_receivable_submission_key",
                            None
                        )

                    if (
                        process_payment_submission
                        and parsed_payment_amount is not None
                    ):

                        st.session_state.receipt_account = (
                            submitted_payment_account
                        )
                        payment_amount = parsed_payment_amount

                        fifo_df = detail_df.copy()

                        fifo_df["_請求日"] = pd.to_datetime(
                            fifo_df["請求日"],
                            errors="coerce"
                        )

                        fifo_df["_表示順"] = range(
                            len(fifo_df)
                        )

                        fifo_df = fifo_df.sort_values(
                            ["_請求日", "_表示順"],
                            na_position="last",
                            kind="stable"
                        )

                        target_candidates = []
                        partial_candidates = []
                        remaining = int(payment_amount)

                        for _, detail in fifo_df.iterrows():

                            target_amount = int(detail["残高"])

                            if target_amount <= 0:
                                continue

                            target_candidates.append({
                                "コード": detail["コード"],
                                "未収ID": detail["未収ID"],
                                "請求日": detail["請求日"],
                                "請求額": detail["請求金額"],
                                "残高": detail["残高"],
                                "消込予定": target_amount,
                                "未収科目": detail["未収科目"],
                                "未収補助": detail["未収補助"],
                                "部門": detail["部門"],
                                "取引先": customer_name,
                                "摘要": detail.get("摘要", "")
                            })

                            if remaining <= 0:
                                continue

                            scheduled_amount = min(
                                target_amount,
                                remaining
                            )

                            if scheduled_amount <= 0:
                                continue

                            partial_candidates.append({
                                "コード": detail["コード"],
                                "未収ID": detail["未収ID"],
                                "請求日": detail["請求日"],
                                "請求額": detail["請求金額"],
                                "残高": detail["残高"],
                                "消込予定": scheduled_amount,
                                "未収科目": detail["未収科目"],
                                "未収補助": detail["未収補助"],
                                "部門": detail["部門"],
                                "取引先": customer_name,
                                "摘要": detail.get("摘要", "")
                            })

                            remaining -= scheduled_amount

                        target_total = sum(
                            item["消込予定"]
                            for item in target_candidates
                        )
                        difference = calculate_receivable_difference(
                            int(payment_amount),
                            target_total
                        )

                        st.session_state[
                            payment_candidates_key
                        ] = {
                            "payment_date": submitted_payment_date,
                            "payment_amount": int(payment_amount),
                            "receipt_account": submitted_payment_account,
                            "items": (
                                partial_candidates
                                if difference < 0
                                else target_candidates
                            ),
                            "target_items": target_candidates,
                            "partial_items": partial_candidates,
                            "target_total": target_total,
                            "difference": difference
                        }
                        st.session_state.pop(
                            "pending_receivable_submission_key",
                            None
                        )

                    candidate_state = st.session_state.get(
                        payment_candidates_key
                    )

                    if candidate_state:

                        candidates = candidate_state["items"]
                        target_candidates = candidate_state.get(
                            "target_items",
                            candidates
                        )
                        partial_candidates = candidate_state.get(
                            "partial_items",
                            candidates
                        )
                        target_total = int(
                            candidate_state.get(
                                "target_total",
                                sum(
                                    item["消込予定"]
                                    for item in candidates
                                )
                            )
                        )
                        difference = int(
                            candidate_state.get("difference", 0)
                        )

                        if target_candidates:

                            st.write("FIFO候補")

                            st.dataframe(
                                pd.DataFrame(target_candidates).drop(
                                    columns=["未収ID"],
                                    errors="ignore"
                                ),
                                use_container_width=True
                            )

                            st.write(
                                "未収対象合計:",
                                target_total
                            )
                            st.write(
                                "差額:",
                                difference
                            )

                            def execute_receivable_settlement(
                                settlement_candidates,
                                difference_account=None,
                                difference_side=None
                            ):

                                selected_accounts = [
                                    candidate_state["receipt_account"]
                                ]
                                if difference_account:
                                    selected_accounts.append(
                                        difference_account
                                    )

                                missing_account_names = [
                                    account_name
                                    for account_name in selected_accounts
                                    if account_name
                                    and not get_account_code(account_name)
                                ]

                                if missing_account_names:
                                    st.warning(
                                        "科目コードを補完できないため処理できません: "
                                        + "、".join(missing_account_names)
                                    )
                                    return

                                journal_rows = build_receivable_journal_rows(
                                    settlement_candidates,
                                    candidate_state["payment_amount"],
                                    candidate_state["receipt_account"],
                                    customer_name,
                                    difference_account,
                                    difference_side
                                )

                                settlement_id = uuid.uuid4().hex

                                generated_receivable_journal = {
                                    "settlement_id": settlement_id,
                                    "settlement_date": candidate_state[
                                        "payment_date"
                                    ],
                                    "rows": journal_rows,
                                    "source_candidates": copy.deepcopy(
                                        settlement_candidates
                                    ),
                                    "customer_name": customer_name,
                                    "created_at": datetime.now().strftime(
                                        "%Y/%m/%d %H:%M"
                                    )
                                }

                                if (
                                    "receivable_generated_journals"
                                    not in st.session_state
                                ):
                                    st.session_state[
                                        "receivable_generated_journals"
                                    ] = []

                                st.session_state[
                                    "receivable_generated_journals"
                                ].append(generated_receivable_journal)

                                try:
                                    apply_receivable_candidates(
                                        settlement_candidates,
                                        candidate_state["payment_date"],
                                        settlement_id
                                    )
                                except Exception:
                                    st.session_state[
                                        "receivable_generated_journals"
                                    ] = [
                                        journal
                                        for journal in st.session_state[
                                            "receivable_generated_journals"
                                        ]
                                        if (
                                            not isinstance(journal, dict)
                                            or journal.get("settlement_id")
                                            != settlement_id
                                        )
                                    ]
                                    raise

                                del st.session_state[
                                    payment_candidates_key
                                ]

                                st.session_state[
                                    "receivable_success"
                                ] = "消込が完了しました"

                                st.rerun()

                            if difference == 0:

                                st.caption(
                                    "表示された候補で未収を消し込みます。"
                                )
                                if st.button(
                                    "消込実行",
                                    key=f"payment_execute_{customer_idx}",
                                    type="primary"
                                ):

                                    try:
                                        execute_receivable_settlement(
                                            candidates
                                        )
                                    except Exception as e:
                                        st.error(str(e))

                            elif difference < 0:

                                shortage_amount = abs(difference)
                                st.warning(
                                    "入金不足があります。差額処理を選択してください。"
                                )
                                st.write(
                                    "入金額:",
                                    candidate_state["payment_amount"]
                                )
                                st.write("不足額:", shortage_amount)

                                default_expense_account = (
                                    "支払手数料"
                                    if "支払手数料" in account_master
                                    else account_master[0]
                                )
                                (
                                    shortage_account_options,
                                    shortage_recommended_accounts,
                                    default_expense_account
                                ) = (
                                    build_receivable_difference_account_options(
                                        records,
                                        customer_name,
                                        target_candidates,
                                        "debit",
                                        default_expense_account
                                    )
                                )
                                shortage_method_key = (
                                    "shortage_method_"
                                    f"{customer_idx}"
                                )
                                shortage_difference_account_key = (
                                    "shortage_difference_account_"
                                    f"{customer_idx}"
                                )
                                shortage_method = st.radio(
                                    "処理方法",
                                    [
                                        "部分消込（残額を未収に残す）",
                                        "差額を科目で処理する"
                                    ],
                                    key=shortage_method_key
                                )

                                shortage_difference_account = None
                                if shortage_method == "差額を科目で処理する":
                                    if (
                                        st.session_state.get(
                                            shortage_difference_account_key
                                        )
                                        not in shortage_account_options
                                    ):
                                        st.session_state[
                                            shortage_difference_account_key
                                        ] = default_expense_account

                                    shortage_difference_account = st.selectbox(
                                        "差額処理科目",
                                        shortage_account_options,
                                        index=shortage_account_options.index(
                                            default_expense_account
                                        ),
                                        key=shortage_difference_account_key,
                                        format_func=(
                                            lambda account,
                                            recommended=shortage_recommended_accounts:
                                            format_recommended_account(
                                                account,
                                                recommended
                                            )
                                        )
                                    )
                                    st.caption(
                                        "選択した借方科目で不足額を処理します。"
                                    )

                                if st.button(
                                    "この内容で処理",
                                    key=f"shortage_execute_{customer_idx}",
                                    type="primary"
                                ):
                                    try:
                                        if (
                                            shortage_method
                                            == "部分消込（残額を未収に残す）"
                                        ):
                                            execute_receivable_settlement(
                                                partial_candidates
                                            )
                                        else:
                                            selected_difference_account = (
                                                st.session_state.get(
                                                    shortage_difference_account_key,
                                                    shortage_difference_account
                                                )
                                            )
                                            execute_receivable_settlement(
                                                target_candidates,
                                                selected_difference_account,
                                                "debit"
                                            )
                                    except Exception as e:
                                        st.error(str(e))

                            else:

                                overpaid_amount = difference
                                st.warning(
                                    "過入金があります。差額処理を選択してください。"
                                )
                                st.write(
                                    "入金額:",
                                    candidate_state["payment_amount"]
                                )
                                st.write("過入金額:", overpaid_amount)

                                default_suspense_account = (
                                    "仮受金"
                                    if "仮受金" in account_master
                                    else account_master[0]
                                )
                                (
                                    overpaid_account_options,
                                    overpaid_recommended_accounts,
                                    default_suspense_account
                                ) = (
                                    build_receivable_difference_account_options(
                                        records,
                                        customer_name,
                                        target_candidates,
                                        "credit",
                                        default_suspense_account
                                    )
                                )
                                overpaid_difference_account_key = (
                                    "overpaid_difference_account_"
                                    f"{customer_idx}"
                                )
                                if (
                                    st.session_state.get(
                                        overpaid_difference_account_key
                                    )
                                    not in overpaid_account_options
                                ):
                                    st.session_state[
                                        overpaid_difference_account_key
                                    ] = default_suspense_account

                                st.write("処理方法:", "差額を科目で処理する")
                                overpaid_difference_account = st.selectbox(
                                    "差額処理科目",
                                    overpaid_account_options,
                                    index=overpaid_account_options.index(
                                        default_suspense_account
                                    ),
                                    key=overpaid_difference_account_key,
                                    format_func=(
                                        lambda account,
                                        recommended=overpaid_recommended_accounts:
                                        format_recommended_account(
                                            account,
                                            recommended
                                        )
                                    )
                                )
                                st.caption(
                                    "選択した貸方科目で過入金額を処理します。"
                                )

                                if st.button(
                                    "この内容で処理",
                                    key=f"overpaid_execute_{customer_idx}",
                                    type="primary"
                                ):
                                    try:
                                        selected_difference_account = (
                                            st.session_state.get(
                                                overpaid_difference_account_key,
                                                overpaid_difference_account
                                            )
                                        )
                                        execute_receivable_settlement(
                                            target_candidates,
                                            selected_difference_account,
                                            "credit"
                                        )
                                    except Exception as e:
                                        st.error(str(e))

                        else:

                            st.info("消込候補がありません")

    with st.expander("消込履歴（直近50件）"):

        history_df = load_receivable_history()

        if history_df.empty:

            st.info("消込履歴がありません")

        else:

            # ファイル保持とは分け、画面では短期確認分だけ表示する
            history_display_df = (
                history_df.tail(50).iloc[::-1]
            )

            st.dataframe(
                history_display_df[
                    ["消込日", "コード", "消込額"]
                ],
                use_container_width=True
            )

    with st.expander("生成仕訳"):

        if "receivable_generated_journals" not in st.session_state:
            st.session_state["receivable_generated_journals"] = []

        legacy_generated_journal = st.session_state.pop(
            "generated_receivable_journal",
            None
        )

        if legacy_generated_journal:
            existing_ids = {
                journal.get("settlement_id")
                for journal in st.session_state[
                    "receivable_generated_journals"
                ]
                if isinstance(journal, dict)
            }

            if (
                isinstance(legacy_generated_journal, dict)
                and legacy_generated_journal.get("settlement_id")
                not in existing_ids
            ):
                st.session_state[
                    "receivable_generated_journals"
                ].append(legacy_generated_journal)

        generated_journals = st.session_state[
            "receivable_generated_journals"
        ]

        if not generated_journals:

            st.info("生成された仕訳はありません")

        else:

            has_unregistered_journal = False

            for journal_idx, generated_journal in enumerate(
                list(generated_journals)
            ):

                if not isinstance(generated_journal, dict):
                    continue

                journal_rows = generated_journal.get("rows", [])
                settlement_id = generated_journal.get("settlement_id")
                settlement_date = generated_journal.get(
                    "settlement_date"
                )
                source_candidates = generated_journal.get(
                    "source_candidates",
                    []
                )
                journal_customer_name = generated_journal.get(
                    "customer_name",
                    ""
                )
                created_at = generated_journal.get("created_at", "")

                if isinstance(journal_rows, dict):
                    journal_rows = [journal_rows]

                journal_registered = False

                if settlement_id is not None:
                    try:
                        journal_registered = (
                            is_receivable_journal_registered(
                                settlement_id
                            )
                        )
                    except Exception:
                        journal_registered = False

                if journal_registered:
                    st.session_state[
                        "receivable_generated_journals"
                    ] = [
                        journal
                        for journal in st.session_state[
                            "receivable_generated_journals"
                        ]
                        if (
                            not isinstance(journal, dict)
                            or journal.get("settlement_id")
                            != settlement_id
                        )
                    ]
                    continue

                has_unregistered_journal = True
                expander_label = (
                    f"{journal_customer_name or '未収消込'}"
                    f" / {settlement_id or 'IDなし'}"
                )

                if created_at:
                    expander_label = f"{created_at} / {expander_label}"

                with st.expander(expander_label):

                    st.dataframe(
                        pd.DataFrame(journal_rows),
                        use_container_width=True
                    )

                    template_diagnostics = [
                        find_receivable_template_match(
                            journal,
                            journal_customer_name,
                            source_candidates
                        )[1]
                        for journal in journal_rows
                    ]

                    st.caption(
                        "DB雛形検索の確認です。表示のみで、CSV出力値や消込処理は変更しません。"
                    )
                    st.dataframe(
                        pd.DataFrame(template_diagnostics),
                        use_container_width=True
                    )

                    missing_account_names = sorted({
                        account_name
                        for journal in journal_rows
                        for account_name in (
                            journal["借方科目"],
                            journal["貸方科目"]
                        )
                        if account_name
                        and not get_account_code(account_name)
                    })

                    if missing_account_names:
                        st.warning(
                            "科目コードを補完できない科目があります: "
                            + "、".join(missing_account_names)
                            + "。account_master.csvを確認してください。"
                        )

                    if settlement_id is None:

                        st.info(
                            "この仕訳候補には消込IDがありません"
                        )

                    else:
                        st.caption(
                            "生成した仕訳をCSV出力対象へ追加します。"
                        )

                    if (
                        settlement_id is not None
                        and st.button(
                            "この仕訳をCSV出力対象へ登録",
                            key=(
                                "register_receivable_"
                                f"{settlement_id}_{journal_idx}"
                            ),
                            type="primary"
                        )
                    ):

                        try:

                            transaction_rows = []

                            for journal in journal_rows:

                                row = build_receivable_transaction_row(
                                    journal,
                                    settlement_date,
                                    settlement_id,
                                    source_candidates,
                                    journal_customer_name
                                )

                                transaction_rows.append(row)

                            if "confirmed" not in st.session_state:
                                st.session_state.confirmed = []

                            st.session_state.confirmed.append(
                                copy.deepcopy(transaction_rows)
                            )

                            mark_receivable_journal_registered(
                                settlement_id
                            )

                            st.session_state[
                                "receivable_generated_journals"
                            ] = [
                                journal
                                for journal in st.session_state[
                                    "receivable_generated_journals"
                                ]
                                if (
                                    not isinstance(journal, dict)
                                    or journal.get("settlement_id")
                                    != settlement_id
                                )
                            ]

                            st.session_state[
                                "receivable_success"
                            ] = (
                                "仕訳を登録し、CSV出力対象に追加しました"
                            )

                            st.cache_data.clear()
                            st.rerun()

                        except Exception as e:

                            st.error(str(e))

            if not has_unregistered_journal:
                st.info("未登録の生成仕訳はありません")

