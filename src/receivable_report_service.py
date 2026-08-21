"""未収消込結果から確認表の行・filename・xlsx bytesを生成する。"""

from __future__ import annotations

import hashlib
import io
import re
from collections.abc import Mapping, Sequence
from typing import Any


RECEIVABLE_CHECK_COLUMNS = [
    "No",
    "入金日",
    "取引先",
    "入金額",
    "消込額",
    "差額",
    "未収内容",
    "生成仕訳",
    "摘要",
]


def _to_int(value: Any) -> int:
    """既存帳票の金額変換と同じ値を返す。"""

    try:
        return int(float(str(value).replace(",", "")))
    except Exception:
        return 0


def format_receivable_check_date(value: Any) -> str:
    if hasattr(value, "strftime"):
        return value.strftime("%Y/%m/%d")

    text = str(value or "").strip()

    if len(text) == 8 and text.isdigit():
        return f"{text[:4]}/{text[4:6]}/{text[6:]}"

    if "-" in text:
        return text.replace("-", "/")

    return text


def format_receivable_check_amount(value: Any) -> str:
    amount = _to_int(value)

    return f"{amount:,}" if amount else "0"


def build_receivable_check_filename(
    generated_journals: Sequence[Mapping[str, Any]] | None,
) -> str:
    journals = [
        journal
        for journal in (generated_journals or [])
        if isinstance(journal, dict)
    ]
    settlement_ids = sorted({
        str(journal.get("settlement_id", "") or "").strip()
        for journal in journals
        if str(journal.get("settlement_id", "") or "").strip()
    })

    settlement_dates = []
    for journal in journals:
        date_digits = re.sub(
            r"\D",
            "",
            format_receivable_check_date(
                journal.get("settlement_date", "")
            )
        )
        if len(date_digits) >= 8:
            settlement_dates.append(date_digits[:8])

    settlement_date = (
        sorted(settlement_dates)[0]
        if settlement_dates
        else "undated"
    )

    if len(settlement_ids) == 1:
        id_part = re.sub(
            r"[^0-9A-Za-z_-]",
            "",
            settlement_ids[0]
        )[:8]
    else:
        id_part = hashlib.sha256(
            "|".join(settlement_ids).encode("utf-8")
        ).hexdigest()[:8]

    if not id_part:
        id_part = "no_id"

    return f"receivable_check_{settlement_date}_{id_part}.xlsx"


def build_receivable_check_rows(
    generated_journals: Sequence[Mapping[str, Any]] | None,
) -> list[dict[str, Any]]:
    rows = []

    for index, generated_journal in enumerate(
        generated_journals or [],
        start=1
    ):
        if not isinstance(generated_journal, dict):
            continue

        journal_rows = generated_journal.get("rows", [])
        source_candidates = generated_journal.get(
            "source_candidates",
            []
        )

        if isinstance(journal_rows, dict):
            journal_rows = [journal_rows]

        target_total = generated_journal.get("target_total")
        if target_total in ("", None):
            target_total = sum(
                _to_int(candidate.get("消込予定"))
                for candidate in source_candidates
                if isinstance(candidate, dict)
            )
        else:
            target_total = _to_int(target_total)

        payment_amount = generated_journal.get("payment_amount")
        if payment_amount in ("", None):
            difference = generated_journal.get("difference")
            if difference not in ("", None):
                payment_amount = target_total + _to_int(difference)
            else:
                payment_amount = sum(
                    _to_int(journal.get("金額"))
                    for journal in journal_rows
                    if isinstance(journal, dict)
                )
        payment_amount = _to_int(payment_amount)

        difference = generated_journal.get("difference")
        if difference in ("", None):
            difference = payment_amount - target_total
        else:
            difference = _to_int(difference)

        receivable_lines = []
        for candidate in source_candidates:
            if not isinstance(candidate, dict):
                continue

            account = str(candidate.get("未収科目", "") or "").strip()
            amount = _to_int(candidate.get("消込予定"))

            if account or amount:
                receivable_lines.append(
                    f"{account} {format_receivable_check_amount(amount)}"
                )

        journal_lines = []
        summaries = []
        for journal in journal_rows:
            if not isinstance(journal, dict):
                continue

            debit = str(journal.get("借方科目", "") or "").strip()
            credit = str(journal.get("貸方科目", "") or "").strip()
            amount = _to_int(journal.get("金額"))
            summary = str(journal.get("摘要", "") or "").strip()

            if debit or credit or amount:
                journal_lines.append(
                    f"{debit} / {credit} "
                    f"{format_receivable_check_amount(amount)}"
                )

            if summary and summary not in summaries:
                summaries.append(summary)

        if not summaries:
            summaries = [
                str(candidate.get("摘要", "") or "").strip()
                for candidate in source_candidates
                if (
                    isinstance(candidate, dict)
                    and str(candidate.get("摘要", "") or "").strip()
                )
            ]

        rows.append({
            "No": index,
            "入金日": format_receivable_check_date(
                generated_journal.get("settlement_date", "")
            ),
            "取引先": generated_journal.get("customer_name", ""),
            "入金額": payment_amount,
            "消込額": target_total,
            "差額": difference,
            "未収内容": "\n".join(receivable_lines),
            "生成仕訳": "\n".join(journal_lines),
            "摘要": "\n".join(summaries),
        })

    return rows


def build_receivable_check_excel(
    check_rows: Sequence[Mapping[str, Any]],
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.properties import PageSetupProperties

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "未収消込確認表"

    headers = RECEIVABLE_CHECK_COLUMNS
    worksheet.append(headers)

    amount_columns = {
        "入金額",
        "消込額",
        "差額",
    }
    wrap_columns = {
        "未収内容",
        "生成仕訳",
        "摘要",
    }
    column_widths = {
        "No": 4,
        "入金日": 11,
        "取引先": 16,
        "入金額": 11,
        "消込額": 11,
        "差額": 10,
        "未収内容": 18,
        "生成仕訳": 22,
        "摘要": 20,
    }

    for row in check_rows:
        worksheet.append([
            row.get(header, "")
            for header in headers
        ])

    thin_side = Side(style="thin", color="B7B7B7")
    cell_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7"
    )

    for column_index, column_name in enumerate(headers, start=1):
        column_letter = worksheet.cell(
            row=1,
            column=column_index
        ).column_letter
        worksheet.column_dimensions[column_letter].width = (
            column_widths.get(column_name, 12)
        )

    for row in worksheet.iter_rows():
        for cell in row:
            column_name = headers[cell.column - 1]
            cell.border = cell_border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=column_name in wrap_columns
            )
            if column_name in amount_columns and cell.row > 1:
                cell.number_format = "#,##0"

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"
    worksheet.print_title_rows = "1:1"
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    if worksheet.sheet_properties.pageSetUpPr is None:
        worksheet.sheet_properties.pageSetUpPr = PageSetupProperties()
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins.left = 0.2
    worksheet.page_margins.right = 0.2
    worksheet.page_margins.top = 0.35
    worksheet.page_margins.bottom = 0.35
    worksheet.page_margins.header = 0.15
    worksheet.page_margins.footer = 0.15

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
