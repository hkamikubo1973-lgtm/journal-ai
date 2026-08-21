"""入力用Excel（簡易仕訳帳・印刷用）を副作用なしで生成する。"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import datetime
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.properties import PageSetupProperties

from columns import (
    COL_CREDIT,
    COL_CREDIT_AMOUNT,
    COL_CREDIT_SUB,
    COL_DATE,
    COL_DEBIT,
    COL_DEBIT_AMOUNT,
    COL_DEBIT_SUB,
    COL_SUMMARY,
)
from journal_export_service import (
    EpsonExportValidationError,
    validate_epson_export_items,
)


INPUT_CSV_COLUMNS = [
    "No",
    "伝票日付",
    "借方科目",
    "借方補助",
    "借方金額",
    "貸方科目",
    "貸方補助",
    "貸方金額",
    "摘要",
    "伝票摘要",
    "区分",
    "注意",
]


class InputExcelValidationError(ValueError):
    """入力用Excelを一部生成せず、request全体を中止する検証エラー。"""


@dataclass(frozen=True)
class InputExcelExport:
    content: bytes
    filename: str
    rows: tuple[dict[str, Any], ...] = ()


def iter_confirmed_journal_rows(confirmed_journals):
    """Streamlitの登録済仕訳を伝票順・伝票内行順で平坦化する。"""

    for item in confirmed_journals:
        if isinstance(item, dict):
            yield item
        elif isinstance(item, list):
            for row in item:
                if isinstance(row, dict):
                    yield row


def build_input_csv_rows(confirmed_journals):
    """既存Streamlit行から手入力・確認用の12列を作成する。"""

    rows = []

    def first_value(journal, *keys):
        for key in keys:
            value = journal.get(key, "")
            if value not in (None, ""):
                return value
        return ""

    for idx, journal in enumerate(
        iter_confirmed_journal_rows(confirmed_journals),
        start=1,
    ):
        debit_account = first_value(
            journal,
            COL_DEBIT,
            "借方科目名",
            "借方科目",
        )
        debit_sub = first_value(
            journal,
            COL_DEBIT_SUB,
            "借方補助科目名",
            "借方補助",
        )
        debit_amount = first_value(
            journal,
            COL_DEBIT_AMOUNT,
            "借方金額",
            "金額",
        )
        debit_sub_code = first_value(
            journal,
            "借方補助コード",
            "借方補助",
        )

        credit_account = first_value(
            journal,
            COL_CREDIT,
            "貸方科目名",
            "貸方科目",
        )
        credit_sub = first_value(
            journal,
            COL_CREDIT_SUB,
            "貸方補助科目名",
            "貸方補助",
        )
        credit_amount = first_value(
            journal,
            COL_CREDIT_AMOUNT,
            "貸方金額",
            "金額",
        )
        credit_sub_code = first_value(
            journal,
            "貸方補助コード",
            "貸方補助",
        )

        description = first_value(journal, COL_SUMMARY, "摘要")
        voucher_description = first_value(journal, "伝票摘要")

        note_items = []
        if (
            journal.get("DB雛形") == "なし"
            or journal.get("db_template_found") is False
        ):
            note_items.append("DB雛形なし")
        if debit_sub and not debit_sub_code:
            note_items.append("借方補助コード未取得")
        if credit_sub and not credit_sub_code:
            note_items.append("貸方補助コード未取得")
        if not voucher_description:
            note_items.append("伝票摘要なし")

        rows.append({
            "No": idx,
            "伝票日付": first_value(journal, COL_DATE, "伝票日付", "日付"),
            "借方科目": debit_account,
            "借方補助": debit_sub,
            "借方金額": debit_amount,
            "貸方科目": credit_account,
            "貸方補助": credit_sub,
            "貸方金額": credit_amount,
            "摘要": description,
            "伝票摘要": voucher_description,
            "区分": first_value(journal, "区分", "source", "処理区分"),
            "注意": " / ".join(note_items),
        })

    return rows


def _input_headers_and_rows(input_data):
    if hasattr(input_data, "columns") and hasattr(input_data, "iterrows"):
        return list(input_data.columns), (
            row for _, row in input_data.iterrows()
        )
    return list(INPUT_CSV_COLUMNS), iter(input_data)


def build_input_journal_excel(input_data):
    """既存Streamlitと同じ内容・書式のxlsx bytesを生成する。"""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "簡易仕訳帳"

    headers, input_rows = _input_headers_and_rows(input_data)
    worksheet.append(headers)

    amount_columns = {"借方金額", "貸方金額"}
    wrap_columns = {"摘要", "伝票摘要", "注意"}
    wrap_column_letters = set()

    for row in input_rows:
        values = []
        for column in headers:
            value = row.get(column, "")
            if column in amount_columns and value not in ("", None):
                try:
                    value = int(str(value).replace(",", ""))
                except ValueError:
                    pass
            values.append(value)
        worksheet.append(values)

    column_widths = {
        "No": 4,
        "伝票日付": 10,
        "借方科目": 12,
        "借方補助": 14,
        "借方金額": 11,
        "貸方科目": 12,
        "貸方補助": 14,
        "貸方金額": 11,
        "摘要": 24,
        "伝票摘要": 18,
        "区分": 8,
        "注意": 18,
    }

    thin_side = Side(style="thin", color="B7B7B7")
    cell_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for column_index, column_name in enumerate(headers, start=1):
        column_letter = worksheet.cell(
            row=1,
            column=column_index,
        ).column_letter
        worksheet.column_dimensions[column_letter].width = column_widths.get(
            column_name,
            12,
        )
        if column_name in wrap_columns:
            wrap_column_letters.add(column_letter)

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = cell_border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(cell.column_letter in wrap_column_letters),
            )
            if headers[cell.column - 1] in amount_columns and cell.row > 1:
                cell.number_format = "#,##0"

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"
    worksheet.print_title_rows = "1:1"
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    if worksheet.sheet_properties.pageSetUpPr is None:
        worksheet.sheet_properties.pageSetUpPr = PageSetupProperties()
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.5
    worksheet.page_margins.bottom = 0.5
    worksheet.page_margins.header = 0.2
    worksheet.page_margins.footer = 0.2

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _validated_print_fields(
    item: Mapping[str, Any],
    item_number: int,
) -> tuple[str, list[str]]:
    print_metadata = item.get("print_metadata")
    if not isinstance(print_metadata, Mapping):
        raise InputExcelValidationError(
            f"{item_number}件目のprint_metadataがありません。"
        )

    print_category = print_metadata.get("print_category")
    if not isinstance(print_category, str):
        raise InputExcelValidationError(
            f"{item_number}件目のprint_categoryは文字列で指定してください。"
        )

    print_warnings = item.get("print_warnings")
    if not isinstance(print_warnings, list) or any(
        not isinstance(warning, str) for warning in print_warnings
    ):
        raise InputExcelValidationError(
            f"{item_number}件目のprint_warningsは文字列配列で指定してください。"
        )

    return print_category, list(print_warnings)


def build_input_excel_rows(
    items: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    """検証済みカートを順序どおり12列の印刷行へ変換する。"""

    if not items:
        raise InputExcelValidationError(
            "入力用Excelの出力対象がありません。"
        )

    try:
        validate_epson_export_items(items)
    except EpsonExportValidationError as error:
        raise InputExcelValidationError(str(error)) from error

    rows = []
    for item_number, item in enumerate(items, start=1):
        prepared = item["prepared_journal"]
        print_category, print_warnings = _validated_print_fields(
            item,
            item_number,
        )
        rows.append({
            "No": item_number,
            "伝票日付": prepared["voucher_date"],
            "借方科目": (
                prepared["debit_account_name"]
                or prepared["debit_account_code"]
            ),
            "借方補助": (
                prepared["debit_sub_name"]
                or prepared["debit_sub_code"]
            ),
            "借方金額": prepared["amount"],
            "貸方科目": (
                prepared["credit_account_name"]
                or prepared["credit_account_code"]
            ),
            "貸方補助": (
                prepared["credit_sub_name"]
                or prepared["credit_sub_code"]
            ),
            "貸方金額": prepared["amount"],
            "摘要": prepared["summary"],
            "伝票摘要": prepared["voucher_summary"],
            "区分": print_category,
            "注意": " / ".join(print_warnings),
        })

    return rows


def export_input_excel(
    items: Sequence[Mapping[str, Any]],
    *,
    export_datetime: datetime | None = None,
) -> InputExcelExport:
    """カートを検証し、保存やDB更新をせずxlsx bytesを返す。"""

    rows = build_input_excel_rows(items)
    content = build_input_journal_excel(rows)
    current_datetime = export_datetime or datetime.now()
    filename = (
        "input_journal_print_"
        f"{current_datetime.strftime('%Y%m%d_%H%M')}.xlsx"
    )
    return InputExcelExport(
        content=content,
        filename=filename,
        rows=tuple(rows),
    )
